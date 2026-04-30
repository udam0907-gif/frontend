from __future__ import annotations

import json
import os
import re
import tempfile
from pathlib import Path
from typing import Any

from app.core.logging import get_logger
from app.services.llm_service import LLMService

logger = get_logger(__name__)

_SYSTEM_PROMPT = """당신은 한국 정부 R&D 과제용 엑셀 견적서/거래명세서 서식 분석 전문가입니다.
주어진 엑셀 셀 구조를 분석하여 「동적 입력 영역」 필드의 정확한 셀 위치(좌표)를 찾아냅니다.

[절대 원칙 — 반드시 준수]
1. cell_map에는 「동적 입력 영역」만 포함한다.
   동적 입력 영역: 매 출력마다 사용자 입력으로 달라지는 값 (수신처, 작성일자, 품목, 수량, 단가, 합계 등).
   정적 라벨(등록번호·전화·총계 라벨 등), 푸터·인사말, vendor 자기 정보(회사명·사업자번호·주소·업태·담당자·대표자)는
   cell_map에 절대 포함하지 않는다. 이 셀들은 원본 양식 그대로 보존된다.
2. 아래 동적 필드에서 null 반환은 사고 직결이다 — 반드시 찾아야 한다:
   recipient_name, issue_date(또는 issue_date_year/month/day), item_name,
   quantity, unit_price, amount, total_amount
3. 셀 주소는 반드시 "B5" 형식(열문자+행번호). 병합 셀은 반드시 좌상단(앵커) 셀 좌표로 표현.
4. 순수 JSON만 반환. 다른 텍스트 절대 불가."""


class XlsxCellMapper:
    """
    XLSX 파일을 Claude API로 분석해서 필드별 셀 좌표를 반환한다.

    결과(cell_map)를 vendor_template_pool.field_map에 저장하면
    이후 문서 생성 시 재분석 없이 바로 사용 가능하다.
    """

    def __init__(self, llm_service: LLMService) -> None:
        self._llm = llm_service

    async def analyze(self, file_path: str) -> dict[str, Any]:
        """
        XLSX/XLS 파일 분석 → 셀 좌표 맵 반환.
        .xls 파일은 자동으로 .xlsx로 변환 후 분석하며 임시 파일은 정리한다.

        반환 구조:
        {
            "cell_map": { ... },
            "raw_response": str,
            "analyzed_rows": int,
        }
        """
        ext = Path(file_path).suffix.lower()
        if ext not in (".xlsx", ".xls"):
            raise ValueError(f"XLSX/XLS 파일만 분석 가능합니다: {ext}")

        _tmp_converted: str | None = None
        try:
            if ext == ".xls":
                _tmp_converted = self.convert_xls_to_xlsx(file_path)
                actual_path = _tmp_converted
                logger.info("xls_converted_to_xlsx", original=file_path, converted=_tmp_converted)
            else:
                actual_path = file_path

            text_map, analyzed_rows = self._xlsx_to_text_map(actual_path)

            user_msg = (
                f"다음은 한국 기업의 견적서/거래명세서 XLSX 파일의 셀 내용입니다 (최대 {analyzed_rows}행 분석).\n\n"
                f"{text_map}\n\n"
                "아래 규칙에 따라 각 필드가 어느 셀에 해당하는지 분석하여 JSON으로 반환하세요.\n\n"
                "【필드 정의 — 동적 입력 핵심 키 (반드시 cell_map에 포함, null 최소화)】\n"
                "- recipient_name: '귀하', '귀중', '수신처', '貴中' 등 수신자가 입력되는 셀. 라벨 옆/아래 빈 셀.\n"
                "- issue_date: 날짜/작성일/견적일이 입력되는 셀. 반드시 찾을 것 — 누락 = 작성일자 빈칸 사고.\n"
                "- item_name: 품목명/제품명/품명이 입력되는 첫 번째 데이터 행 셀. 헤더 행 아님.\n"
                "- spec: 규격/사양이 입력되는 데이터 행 셀 (없으면 null).\n"
                "- unit: 단위(kg, EA 등)가 입력/고정되는 셀 (없으면 null).\n"
                "- quantity: 수량/QTY가 입력되는 데이터 행 셀. 헤더('수량') 아님.\n"
                "- unit_price: 단가/UNIT PRICE가 입력되는 데이터 행 셀.\n"
                "- amount: 합계/금액/AMOUNT가 입력되는 데이터 행 셀.\n"
                "- total_amount: 총합계/합계(VAT 포함/제외 무관)가 표시되는 셀.\n\n"
                "【추가 필드 — 동적 입력이면 포함】\n"
                "- doc_number: 문서번호/견적번호가 입력되는 셀.\n"
                "- total_amount_korean: 합계 한글 표기 셀 (있으면 포함).\n"
                "- note / remark: 비고란이 입력되는 셀 (있으면 포함).\n"
                "- tax_amount: 부가세 금액 셀.\n"
                "- supply_amount: 공급가액(VAT 제외) 셀.\n\n"
                "【명시적 제외 항목 — cell_map 포함 금지】\n"
                "아래 항목은 원본 양식 그대로 보존되므로 cell_map에 절대 포함하지 않는다:\n"
                "- 정적 라벨: 「등록번호 :」, 「전화 :」, 「주소 :」, 「업태 :」, 「총계」, 「공급가합계」 등 라벨 텍스트\n"
                "- 푸터·인사말: 「* 공장도착도 공급가임.」, 「- 오늘도 좋은 하루 되십시오 -」, 주의사항 문구 등\n"
                "- vendor 자기 정보: 회사명, 사업자등록번호, 주소, 업태, 종목, 전화, 팩스, 담당자, 대표자 등\n"
                "  (vendor 양식에 이미 정확히 기재돼 있으므로 보존 정책 적용)\n\n"
                "【주의사항 — 반드시 준수】\n"
                "1. 헤더(라벨) 행과 데이터 행을 반드시 구분할 것.\n"
                "   - '품목명', '제품명', '품명', 'COMMODITY', '수량', '단가', '금액' 등 라벨이 있는 셀은 헤더 행.\n"
                "   - item_name/quantity/unit_price/amount는 헤더 바로 아래 첫 번째 데이터 입력 행의 셀 주소.\n"
                "   - 예: 헤더가 A11행이면 item_name = 'A12' (데이터 첫 행)\n"
                "2. quantity, unit_price, amount 중 null이면 품목 테이블 전체를 다시 확인. 수량/단가/금액 열이 반드시 존재.\n"
                "3. 병합된 셀은 반드시 좌상단(앵커) 셀 좌표로 표현. (예: B2:E2 병합 → 'B2')\n"
                "4. 반드시 JSON만 반환. 다른 텍스트 없음.\n\n"
                "【품목 테이블 처리 — 다중 행 지원】\n"
                "견적서/거래명세서에 품목 목록(line_items) 표가 있으면, cell_map 안에 '_meta' 키를 추가로 반환:\n"
                '  "_meta": {"items_table": {"start_row": <첫 번째 데이터 행 번호(정수)>, '
                '"columns": {"item_name": "<열문자>", "spec": "<열문자>", "unit": "<열문자>", '
                '"quantity": "<열문자>", "unit_price": "<열문자>", "amount": "<열문자>"}, '
                '"max_rows": <품목 최대 입력 가능 행 수(정수)>}}\n'
                "columns 안의 열문자가 없으면 null. 품목 표가 없는 단일 항목 양식이면 '_meta'를 생략.\n\n"
                "【작성일자 처리 — 분리 셀 지원】\n"
                "issue_date는 반드시 cell_map에 포함. 누락 = 작성일자 빈칸 사고 직결.\n"
                "작성일자 셀이 년/월/일 분리 구조(예: F3=년도, G3=월, H3=일)이면 3개 키로 반환:\n"
                '  "issue_date_year": "F3", "issue_date_month": "G3", "issue_date_day": "H3"\n'
                "분리되지 않고 단일 셀이면 'issue_date' 키 하나만. 둘 중 하나만 (혼합 금지).\n\n"
                "【반환 형식 — 반드시 이 구조 그대로】\n"
                "{\n"
                '  "sheet_name": "시트명",\n'
                '  "cell_map": {\n'
                '    "recipient_name": "B4",\n'
                '    "issue_date": "A3",\n'
                '    "item_name": "A17",\n'
                '    "spec": "B17",\n'
                '    "unit": "C17",\n'
                '    "quantity": "D17",\n'
                '    "unit_price": "E17",\n'
                '    "amount": "F17",\n'
                '    "total_amount": "F30",\n'
                '    "doc_number": "G3",\n'
                '    "_meta": {"items_table": {"start_row": 17, "columns": {"item_name": "A", "spec": "B", "unit": "C", "quantity": "D", "unit_price": "E", "amount": "F"}, "max_rows": 10}}\n'
                "  }\n"
                "}\n"
                "절대 cell_map 안에 cell_map을 중첩하지 말 것. sheet_name은 최상위에만 위치.\n"
                "날짜 분리 셀이면 issue_date 대신 issue_date_year/issue_date_month/issue_date_day 사용.\n"
                "JSON 외 다른 텍스트 없이 순수 JSON만 반환.\n"
            )

            response = await self._llm.complete(
                system_prompt=_SYSTEM_PROMPT,
                user_message=user_msg,
                prompt_version="xlsx-cell-mapper-1.0",
                cache_system=True,
            )

            cell_map = self._parse_response(response.content)

            logger.info(
                "xlsx_cell_map_analyzed",
                file=file_path,
                analyzed_rows=analyzed_rows,
                token_usage=response.token_usage,
            )

            return {
                "cell_map": cell_map,
                "raw_response": response.content,
                "analyzed_rows": analyzed_rows,
            }

        finally:
            if _tmp_converted:
                try:
                    os.unlink(_tmp_converted)
                except Exception:
                    pass

    @staticmethod
    def convert_xls_to_xlsx(xls_path: str) -> str:
        """
        .xls 파일을 .xlsx로 변환하여 임시 파일 경로를 반환한다.
        원본 파일은 수정하지 않는다. 셀 매핑 분석 전용.
        """
        import xlrd
        import openpyxl

        wb_xls = xlrd.open_workbook(xls_path)
        wb_xlsx = openpyxl.Workbook()
        wb_xlsx.remove(wb_xlsx.active)

        for sheet_idx in range(wb_xls.nsheets):
            ws_xls = wb_xls.sheet_by_index(sheet_idx)
            ws_xlsx = wb_xlsx.create_sheet(title=ws_xls.name)

            for row_idx in range(ws_xls.nrows):
                for col_idx in range(ws_xls.ncols):
                    cell = ws_xls.cell(row_idx, col_idx)
                    # xlrd cell type: 0=empty, 1=text, 2=number, 3=date, 4=bool, 5=error
                    if cell.ctype == 2:
                        value = int(cell.value) if cell.value == int(cell.value) else cell.value
                    elif cell.ctype == 3:
                        try:
                            value = xlrd.xldate_as_datetime(cell.value, wb_xls.datemode)
                        except Exception:
                            value = cell.value
                    elif cell.ctype == 4:
                        value = bool(cell.value)
                    elif cell.ctype in (0, 5):
                        value = None
                    else:
                        value = cell.value

                    if value is not None:
                        ws_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=value)

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb_xlsx.save(tmp.name)
        return tmp.name

    def _xlsx_to_text_map(
        self, file_path: str, max_rows: int = 60, max_cols: int = 60
    ) -> tuple[str, int]:
        """
        XLSX를 "A1=값, B1=값, ..." 형태의 텍스트로 변환.
        빈 셀은 제외해서 토큰 절약. 병합 셀 정보도 포함.
        """
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl이 설치되지 않았습니다.")

        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        sheet = wb.active
        sheet_name = sheet.title

        lines: list[str] = [f"[시트명: {sheet_name}]"]

        if sheet.merged_cells.ranges:
            merge_info = ", ".join(str(r) for r in list(sheet.merged_cells.ranges)[:10])
            lines.append(f"[병합셀: {merge_info}]")

        actual_rows = min(sheet.max_row or max_rows, max_rows)
        actual_cols = min(sheet.max_column or max_cols, max_cols)

        for row_idx in range(1, actual_rows + 1):
            cells_in_row: list[str] = []
            for col_idx in range(1, actual_cols + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None or str(val).strip() == "":
                    continue
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                val_str = str(val)[:30]
                cells_in_row.append(f"{col_letter}{row_idx}={val_str}")
            if cells_in_row:
                lines.append("  ".join(cells_in_row))

        wb.close()
        return "\n".join(lines), actual_rows

    def _parse_response(self, content: str) -> dict[str, Any]:
        """
        Claude 응답에서 JSON을 추출하고 중첩 cell_map 구조를 정규화한다.

        지원 패턴:
          패턴 A (정상): {"sheet_name": "X", "cell_map": {"item_name": "A9", ...}}
          패턴 B (중첩): {"cell_map": {"sheet_name": "X", "cell_map": {"item_name": "A9", ...}}}
          패턴 C (sheet 없음): {"cell_map": {"item_name": "A9", ...}}

        반환: {"item_name": "A9", ..., "sheet_name": "X"} 형태의 플랫 딕셔너리
        """
        # 마크다운 코드펜스 제거 (LLM이 ```json ... ``` 로 감싸는 경우)
        stripped = re.sub(r"^```[a-zA-Z]*\s*", "", content.strip())
        stripped = re.sub(r"\s*```\s*$", "", stripped)

        match = re.search(r"\{.*\}", stripped, re.DOTALL)
        if not match:
            logger.warning("xlsx_cell_map_parse_failed", preview=content[:300])
            return {}

        try:
            parsed = json.loads(match.group())
        except json.JSONDecodeError:
            logger.warning("xlsx_cell_map_parse_failed", preview=content[:300])
            return {}

        raw_cell_map = parsed.get("cell_map", {})

        # 중첩 감지: raw_cell_map 안에 또 "cell_map" 키가 있으면 중첩 구조 (패턴 B)
        if isinstance(raw_cell_map, dict) and "cell_map" in raw_cell_map:
            sheet_name = raw_cell_map.get("sheet_name") or parsed.get("sheet_name")
            cell_map = raw_cell_map.get("cell_map", {})
        else:
            sheet_name = parsed.get("sheet_name")
            cell_map = raw_cell_map

        if not isinstance(cell_map, dict):
            logger.warning("xlsx_cell_map_unexpected_type", type=type(cell_map).__name__)
            return {}

        if sheet_name:
            cell_map["sheet_name"] = sheet_name

        return cell_map
