from __future__ import annotations

import shutil
import uuid
from pathlib import Path
from typing import Any

from openpyxl.styles import Font
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from app.config import settings
from app.core.exceptions import DocumentGenerationError, MappingNotFoundError
from app.core.logging import get_logger
from app.services.xlsx_cell_mapper import _is_label_cell

logger = get_logger(__name__)


def _safe_set_cell(
    ws: Any,
    anchor_addr: str,
    value: Any,
    label_skip_log: list[str] | None = None,
    formula_skip_log: list[str] | None = None,
) -> bool:
    """
    cell_map 매핑 결과 셀에 값 박기 직전 두 가지 가드 적용:

    1) 라벨 가드 — mapper가 매핑한 셀이 _anchor()로 병합 좌상단으로 보정됐을 때,
       그 좌상단이 라벨(예: B21=TOTAL, B12=년)이면 박지 않고 보존.
    2) 수식 가드 — 양식의 자동계산 수식(예: N15=`=J15*L15`)이 있는 셀에는
       하드코딩 값을 박지 않고 수식 보존 → Excel이 자동 재계산.

    openpyxl의 load_workbook 기본값(data_only=False)에서 수식은 `=`로 시작하는
    문자열로 노출되므로 startswith('=')로 판정 가능.

    Returns: True=박음, False=라벨/수식이라 거부
    """
    cell = ws[anchor_addr]
    existing = cell.value
    # 1) 라벨 가드
    if existing and isinstance(existing, str) and _is_label_cell(existing):
        if label_skip_log is not None:
            label_skip_log.append(f"{anchor_addr}='{existing}' (label, skip)")
        return False
    # 2) 수식 가드
    if existing and isinstance(existing, str) and existing.startswith("="):
        if formula_skip_log is not None:
            formula_skip_log.append(f"{anchor_addr}='{existing}' (formula, skip)")
        return False
    _set_cell(ws, anchor_addr, value)
    return True


def _set_cell(ws: Any, anchor_addr: str, value: Any) -> None:
    """
    셀에 값을 박고 폰트 색만 검정으로 리셋.
    family/size/bold/italic은 보존하여 양식 미관 유지.

    이유: mapper가 이전에 라벨/placeholder 셀을 가리켰던 경우 그 셀의 폰트가
    보라(라벨용)/회색(placeholder용)인데, 데이터를 그대로 넣으면 색이 transfer됨.
    값 박을 때만 검정으로 리셋해 데이터 가독성 보장.
    """
    cell = ws[anchor_addr]
    cell.value = value
    if cell.font:
        cell.font = Font(
            name=cell.font.name,
            size=cell.font.size,
            bold=cell.font.bold,
            italic=cell.font.italic,
            color="FF000000",
        )


def _is_digit_breakdown_layout(ws: Any, start_ref: str) -> bool:
    """
    start_ref 셀의 한 행 위가 자릿수 라벨(억/천/백/십/만/일) 6개 이상이면
    합계금액 자릿수 분리 양식으로 판정.

    옵토마린 양식 예: E12-M12 라벨(억/천/백/십/만/천/백/십/일), E13 데이터 시작.
    """
    col_letter = ''.join(c for c in start_ref if c.isalpha())
    row_str = ''.join(c for c in start_ref if c.isdigit())
    if not col_letter or not row_str:
        return False
    row = int(row_str)
    if row <= 1:
        return False
    label_row = row - 1
    DIGIT_LABELS = {"억", "천", "백", "십", "만", "일"}
    start_col = column_index_from_string(col_letter)
    label_count = 0
    for i in range(9):
        ref = f"{get_column_letter(start_col + i)}{label_row}"
        try:
            v = ws[ref].value
        except Exception:
            continue
        if v and str(v).strip() in DIGIT_LABELS:
            label_count += 1
    return label_count >= 6


def _fill_amount_digits(
    ws: Any, total_amount: int | None, start_cell: str, num_digits: int = 9
) -> list[str]:
    """
    합계금액을 자릿수별로 분리해서 num_digits개 셀에 가로로 박음.

    양식 예 (옵토마린 거래명세서): 행12=라벨(억/천/백/십/만/천/백/십/일),
    행13=값. _fill_amount_digits(ws, 275000, "E13") → E13:M13 의 채움 결과는
    아래 정책 참조.

    0 표시 정책 (γ — 가장 큰 유효 자리부터 끝까지 박음):
      - 9자리 zero-pad 문자열에서 첫 비-0 자리(first_nonzero) 찾는다.
      - first_nonzero 이전 자리는 빈칸(None) — 양식 미관 우선.
      - first_nonzero 이후 자리는 0 포함 모두 정수로 박기 — 가독성/회계 관행.
      - 예) 275,000 → "000275000" → E,F,G=빈칸, H=2, I=7, J=5, K=0, L=0, M=0
      - 예) 5 → "000000005" → M=5, 나머지 빈칸
      - 예) 0 → fallback: M=0, 나머지 빈칸 (마지막 자리 0 표시)
      - 예) 1억 → "100000000" → E~M 모두 박힘 (1,0,0,0,0,0,0,0,0)

    Args:
        ws: 워크시트
        total_amount: 합계 (정수). None/음수면 아무것도 안 함.
        start_cell: 가장 큰 자릿수 셀 좌표 (예: "E13").
        num_digits: 자릿수 개수 (기본 9).

    Returns:
        시도한 셀 좌표 리스트 (None 할당 셀도 포함).
    """
    if total_amount is None or total_amount < 0:
        return []

    col_letter = ''.join(c for c in start_cell if c.isalpha())
    if not col_letter:
        return []
    row_str = ''.join(c for c in start_cell if c.isdigit())
    if not row_str:
        return []
    row = int(row_str)
    start_col = column_index_from_string(col_letter)

    # num_digits 자리 0-padded 문자열 (양식 한계 초과 시 하위 num_digits 자리만)
    s = str(int(total_amount)).zfill(num_digits)
    if len(s) > num_digits:
        s = s[-num_digits:]

    # γ 정책: 첫 비-0 자리 찾기. 0원이면 마지막 자리에 0 표시 (fallback).
    first_nonzero: int
    try:
        first_nonzero = next(i for i, d in enumerate(s) if d != "0")
    except StopIteration:
        first_nonzero = len(s) - 1  # 0원 — 마지막 자리에만 0

    filled: list[str] = []
    for i, digit in enumerate(s):
        ref = f"{get_column_letter(start_col + i)}{row}"
        if i < first_nonzero:
            ws[ref].value = None  # 가장 큰 유효 자리 이전은 빈칸
        else:
            ws[ref].value = int(digit)  # 0 포함 모두 박음
        filled.append(ref)
    return filled


def _parse_date(date_str: str):
    """날짜 문자열(YYYY-MM-DD 등)을 date 객체로 파싱. 실패 시 None 반환."""
    import re
    from datetime import date
    parts = re.split(r"[-./\s]+", date_str.strip())
    parts = [p for p in parts if p.isdigit()]
    if len(parts) >= 3:
        try:
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
        except (ValueError, OverflowError):
            pass
    return None


def _anchor(ws: Any, cell_addr: str) -> str:
    """
    cell_addr이 병합 영역에 속하면 그 영역의 좌상단(앵커) 좌표를 반환.
    아니면 입력된 cell_addr 그대로 반환.

    openpyxl은 병합 영역의 비-앵커 좌표에 값을 쓰면 무시한다.
    이 헬퍼로 cell_map이 어느 좌표를 반환하든 안전하게 쓰기 가능.
    """
    from openpyxl.utils.cell import (
        coordinate_from_string,
        column_index_from_string,
        get_column_letter,
    )

    try:
        col_letter, row = coordinate_from_string(cell_addr)
        col_idx = column_index_from_string(col_letter)
    except Exception:
        return cell_addr

    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col_idx <= merged_range.max_col):
            anchor_col = get_column_letter(merged_range.min_col)
            anchor_row = merged_range.min_row
            return f"{anchor_col}{anchor_row}"
    return cell_addr


class XlsxDocumentFiller:
    """
    cell_map 기반으로 XLSX 템플릿에 값을 채워 새 파일을 생성한다.

    cell_map 구조 (플랫):
    {
        "sheet_name": "시트명",
        "item_name": "A16",
        "spec": "C16",
        "quantity": "E15",
        "unit_price": "F15",
        "amount": "G15",
        "recipient_name": "A4",
        "issue_date": "B9",
        "total_amount": "G26",
        "doc_number": "B8",
        ... 기타 필드
    }

    context 키 → cell_map 키 매핑으로 셀에 값을 채운다.
    """

    # context 키 → cell_map 키 매핑 테이블
    # context에서 값을 가져와서 cell_map의 해당 셀 주소에 쓴다
    FIELD_ALIASES: dict[str, list[str]] = {
        # 수신자/귀하
        "recipient_name": ["recipient_name", "recipient", "귀하", "수신처"],
        # 공급받는자(우리 회사) 영역
        "recipient_business_number": ["recipient_business_number", "our_company_registration_number"],
        "recipient_company_name": ["recipient_company_name", "our_company_name"],
        "recipient_representative": ["recipient_representative", "our_company_representative"],
        "recipient_address": ["recipient_address", "our_company_address"],
        "recipient_business_type": ["recipient_business_type", "our_company_business_type"],
        "recipient_business_item": ["recipient_business_item", "our_company_business_item"],
        # 공급받는자 메일/전화/팩스 — context엔 our_company_* 로 들어옴 (company_setting에서)
        "recipient_email": ["recipient_email", "our_company_email", "company_email"],
        "recipient_phone": ["recipient_phone", "our_company_phone", "company_phone"],
        "recipient_fax":   ["recipient_fax", "our_company_fax", "company_fax"],
        # 날짜
        "issue_date": ["issue_date", "execution_date", "expense_date", "작성일", "견적일"],
        # 품목명
        "item_name": ["item_name", "product_name", "title"],
        # 규격
        "spec": ["spec", "specification"],
        # 수량
        "quantity": ["quantity", "qty"],
        # 단가
        "unit_price": ["unit_price", "price"],
        # 합계/금액
        "amount": ["amount", "total"],
        # 총합계
        "total_amount": ["total_amount", "amount"],
        # 공급가액 합계 (구 supply_amount 호환)
        "total_supply": ["total_supply", "supply_amount"],
        "supply_amount": ["supply_amount", "total_supply"],
        # 세액 합계 (구 tax_amount 호환)
        "total_tax": ["total_tax", "tax_amount"],
        "tax_amount": ["tax_amount", "total_tax"],
        # 문서번호
        "doc_number": ["doc_number", "document_number"],
        # 업체명 (공급자)
        "company_name": ["company_name", "supplier_name", "vendor_name"],
        # 사업자번호
        "registration_number": ["registration_number", "vendor_business_number",
                                 "vendor_registration", "business_number",
                                 "vendor_business_no"],
        # 담당자/연락처
        "contact": ["contact", "vendor_contact"],
        # 예산항목
        "budget_item": ["budget_item"],
        # 비고
        "remark": ["remark", "note"],
    }

    def __init__(self) -> None:
        self._output_base = Path(settings.storage_documents_path)
        self._output_base.mkdir(parents=True, exist_ok=True)

    def fill(
        self,
        template_path: str,
        field_map: dict[str, Any],
        context: dict[str, Any],
        expense_item_id: str,
    ) -> str:
        """
        템플릿을 복사 후 context 값으로 셀을 채운다.
        반환값: 생성된 파일 경로
        """
        try:
            import openpyxl
        except ImportError:
            raise DocumentGenerationError("openpyxl이 설치되지 않았습니다.")

        cell_map: dict = field_map.get("_cell_map", {})
        if not cell_map:
            raise MappingNotFoundError(
                "셀 매핑(cell_map)이 없습니다. "
                "업체 관리에서 양식 파일을 다시 업로드하여 매핑을 완료하세요."
            )

        # 1. 템플릿 복사 (원본 보존)
        # .xls이면 .xlsx로 변환하여 저장 (openpyxl은 .xls 미지원)
        src_ext = Path(template_path).suffix.lower()
        output_filename = f"{expense_item_id}_{uuid.uuid4().hex[:8]}.xlsx"
        output_path = str(self._output_base / output_filename)

        if src_ext == ".xls":
            import os as _os
            from app.services.xlsx_cell_mapper import XlsxCellMapper
            converted = XlsxCellMapper.convert_xls_to_xlsx(template_path)
            try:
                shutil.copy2(converted, output_path)
            finally:
                try:
                    _os.unlink(converted)
                except Exception:
                    pass
        else:
            shutil.copy2(template_path, output_path)

        # 2. 복사본 열기
        wb = openpyxl.load_workbook(output_path)
        sheet_name = cell_map.get("sheet_name")
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        # 3. 플랫 cell_map으로 셀 채움
        written: list[str] = []
        skipped: list[str] = []
        self._fill_flat(ws, cell_map, context, written, skipped)

        # 3.5. (v5) 비주력 시트는 보존 — workbook 메타/drawing 관계/named range 깨짐 방지
        # 데이터 누출 위험은 _fill_flat이 매칭된 단일 ws에만 쓴다는 사실 + Gate 2 시트
        # 화이트리스트 가드로 차단. 매칭된 시트만 활성 탭으로 설정해 사용자가 파일 열 때
        # 거래명세서 시트가 먼저 보이게 한다.
        if sheet_name and sheet_name in wb.sheetnames:
            wb.active = wb.sheetnames.index(sheet_name)
            if len(wb.sheetnames) > 1:
                _other = [s for s in wb.sheetnames if s != sheet_name]
                logger.info(
                    "xlsx_secondary_sheets_preserved",
                    target=sheet_name,
                    other_sheets=_other,
                    other_count=len(_other),
                )

        # 4. 저장
        wb.save(output_path)
        wb.close()

        logger.info(
            "xlsx_document_filled",
            output=output_path,
            items=len(written),
            written=written,
            skipped=skipped if skipped else None,
        )
        return output_path

    def _fill_flat(
        self,
        ws: Any,
        cell_map: dict,
        context: dict,
        written: list,
        skipped: list,
    ) -> None:
        """
        플랫 cell_map의 모든 키를 순회하며 context에서 값을 찾아 셀에 쓴다.

        매핑 방식:
        1. cell_map 키와 동일한 context 키 직접 매핑
        2. FIELD_ALIASES 역방향 매핑 (context 키 → cell_map 키)
        3. cell_map 키에 대해 FIELD_ALIASES로 context에서 값 탐색

        확장:
        - _meta.items_table 존재 시 line_items 배열을 다중 행으로 입력
        - issue_date_year/month/day 존재 시 날짜를 분리 셀에 입력
        """
        skip_keys = {
            "sheet_name", "_cell_map", "_mapping_status", "_meta",
            "issue_date_year", "issue_date_month", "issue_date_day",
        }

        # ── 다중 행 라인아이템 처리 ──────────────────────────────────────
        items_meta: dict | None = None
        if isinstance(cell_map.get("_meta"), dict):
            items_meta = cell_map["_meta"].get("items_table")
        if items_meta and context.get("line_items"):
            start_row = items_meta.get("start_row")
            columns: dict = items_meta.get("columns", {})

            # ── 라인별 날짜 자동 주입 (월/일/날짜/월일 컬럼 지원) ────────
            if start_row and columns:
                _line_date_keys = {"month", "day", "date", "month_day"}
                if any(k in columns for k in _line_date_keys):
                    _raw_d = (
                        context.get("expense_date")
                        or context.get("issue_date")
                        or context.get("execution_date")
                    )
                    _d_obj = None
                    if _raw_d is not None:
                        if hasattr(_raw_d, "year"):
                            _d_obj = _raw_d
                        else:
                            _d_obj = _parse_date(str(_raw_d))
                    if _d_obj:
                        for _it in context["line_items"]:
                            _it.setdefault("month", _d_obj.month)
                            _it.setdefault("day", _d_obj.day)
                            _it.setdefault("date", f"{_d_obj.month}/{_d_obj.day}")
                            _it.setdefault("month_day", f"{_d_obj.month} {_d_obj.day}")

            if start_row and columns:
                # spec 컬럼이 없거나 item_name과 같은 병합 셀을 가리키면 item_name에 합쳐서 쓴다.
                # 예: 대신테크젠(spec=None), 민서정밀(spec=None), 선양(spec='D', item_name='C', C9:D9 병합)
                _spec_col = columns.get("spec")
                _name_col = columns.get("item_name")
                _merge_spec_into_name = False
                if _name_col:
                    if not _spec_col:
                        _merge_spec_into_name = True
                    else:
                        try:
                            _probe_row = int(start_row)
                            if _anchor(ws, f"{_spec_col}{_probe_row}") == _anchor(ws, f"{_name_col}{_probe_row}"):
                                _merge_spec_into_name = True
                        except Exception:
                            pass

                for idx, item in enumerate(context["line_items"]):
                    row = int(start_row) + idx

                    # spec을 item_name에 합칠 경우 병합 값 미리 계산
                    _merged_item_name: str | None = None
                    if _merge_spec_into_name and _name_col:
                        _n = item.get("item_name") or ""
                        _s = item.get("spec") or ""
                        _merged_item_name = f"{_n} ({_s})" if (_n and _s) else (_n or _s or None)

                    # 행별 앵커 충돌 방지 — 먼저 쓴 필드가 해당 앵커를 선점
                    _row_claimed_anchors: dict[str, str] = {}

                    for field_key, col_letter in columns.items():
                        # col_letter가 None이면 쓸 위치가 없음 — skip
                        if not col_letter:
                            skipped.append(f"line_items[{idx}].{field_key}: col=None")
                            continue

                        val = item.get(field_key)
                        if val is None and field_key == "unit_price":
                            val = item.get("price")

                        # spec을 item_name에 합치는 경우
                        if _merge_spec_into_name:
                            if field_key == "item_name" and _merged_item_name is not None:
                                val = _merged_item_name
                            elif field_key == "spec":
                                skipped.append(f"line_items[{idx}].spec: merged into item_name")
                                continue

                        if val is None or val == "":
                            continue
                        try:
                            from decimal import Decimal
                            if isinstance(val, Decimal):
                                val = int(val) if val == val.to_integral_value() else float(val)
                        except Exception:
                            pass
                        try:
                            cell_addr = f"{col_letter}{row}"
                            anchor_addr = _anchor(ws, cell_addr)
                            # 같은 행에서 이미 이 앵커에 값을 쓴 필드가 있으면 skip
                            if anchor_addr in _row_claimed_anchors:
                                skipped.append(
                                    f"line_items[{idx}].{field_key}({cell_addr}): "
                                    f"anchor {anchor_addr} already claimed by {_row_claimed_anchors[anchor_addr]}"
                                )
                                continue
                            if _safe_set_cell(ws, anchor_addr, val):
                                _row_claimed_anchors[anchor_addr] = field_key
                                written.append(f"line_items[{idx}].{field_key}={col_letter}{row}")
                            else:
                                skipped.append(
                                    f"line_items[{idx}].{field_key}({col_letter}{row}→{anchor_addr}): "
                                    f"label/formula protected"
                                )
                        except Exception as e:
                            skipped.append(f"line_items[{idx}].{field_key}({col_letter}{row}): {e}")
                # items_table 컬럼 키는 단일 셀 루프에서 제외
                for col_key in columns:
                    skip_keys.add(col_key)

        # ── 날짜 분리 셀 처리 ────────────────────────────────────────────
        year_cell = cell_map.get("issue_date_year")
        month_cell = cell_map.get("issue_date_month")
        day_cell = cell_map.get("issue_date_day")
        if year_cell or month_cell or day_cell:
            raw_date = (
                context.get("expense_date")
                or context.get("issue_date")
                or context.get("execution_date")
            )
            date_obj = None
            if raw_date is not None:
                if hasattr(raw_date, "year"):
                    date_obj = raw_date
                else:
                    date_obj = _parse_date(str(raw_date))
            if date_obj is not None:
                if year_cell:
                    try:
                        a = _anchor(ws, year_cell)
                        if _safe_set_cell(ws, a, date_obj.year):
                            written.append(f"issue_date_year={year_cell}")
                        else:
                            skipped.append(f"issue_date_year({year_cell}→{a}): label cell protected")
                    except Exception as e:
                        skipped.append(f"issue_date_year({year_cell}): {e}")
                if month_cell:
                    try:
                        a = _anchor(ws, month_cell)
                        if _safe_set_cell(ws, a, date_obj.month):
                            written.append(f"issue_date_month={month_cell}")
                        else:
                            skipped.append(f"issue_date_month({month_cell}→{a}): label cell protected")
                    except Exception as e:
                        skipped.append(f"issue_date_month({month_cell}): {e}")
                if day_cell:
                    try:
                        a = _anchor(ws, day_cell)
                        if _safe_set_cell(ws, a, date_obj.day):
                            written.append(f"issue_date_day={day_cell}")
                        else:
                            skipped.append(f"issue_date_day({day_cell}→{a}): label cell protected")
                    except Exception as e:
                        skipped.append(f"issue_date_day({day_cell}): {e}")
            # 분리 처리 여부와 무관하게 issue_date 단일 키는 단일 셀 루프에서 제외
            skip_keys.add("issue_date")

        # ── 합계금액 자릿수 분리 처리 ────────────────────────────────────
        # 우선순위 1: cell_map.amount_digit_breakdown_start (사용자 명시 marker)
        # 우선순위 2: total_amount 셀의 위 행이 자릿수 라벨(억/천/백/십/만/일)이면 자동 감지
        digit_start = cell_map.get("amount_digit_breakdown_start")
        if not digit_start:
            ta_cell = cell_map.get("total_amount")
            if (
                ta_cell
                and isinstance(ta_cell, str)
                and _is_digit_breakdown_layout(ws, ta_cell)
            ):
                digit_start = ta_cell
        if digit_start and isinstance(digit_start, str):
            total_amount_value = (
                context.get("total_amount")
                or context.get("amount")
                or 0
            )
            try:
                amt = int(total_amount_value)
            except (TypeError, ValueError):
                amt = 0
            if amt > 0:
                filled = _fill_amount_digits(ws, amt, digit_start)
                if filled:
                    written.append(f"amount_digits[{digit_start}]={len(filled)}cells")
                    # 자릿수로 분리됐으니 total_amount 단일 셀 박기 스킵
                    skip_keys.add("total_amount")
            # marker 키는 단일 셀 루프에서 항상 제외 (값이 셀 좌표라 직접 박으면 안 됨)
            skip_keys.add("amount_digit_breakdown_start")

        # ── 단일 셀 1:1 매핑 (기존 로직 그대로) ─────────────────────────
        for cell_key, cell_addr in cell_map.items():
            if cell_key in skip_keys:
                continue
            if not isinstance(cell_addr, str) or not cell_addr:
                continue

            # 값 탐색 순서:
            # 1. context에 cell_key와 동일한 키가 있으면 직접 사용
            value = context.get(cell_key)

            # 2. 없으면 FIELD_ALIASES에서 cell_key에 해당하는 context 키들을 탐색
            if value is None and cell_key in self.FIELD_ALIASES:
                for alias in self.FIELD_ALIASES[cell_key]:
                    value = context.get(alias)
                    if value is not None:
                        break

            # 3. 역방향: cell_key가 어떤 alias에 포함된 경우
            #    (ex: cell_map에 "supplier_name"이 있고 FIELD_ALIASES["company_name"]에
            #    "supplier_name"이 있으면 company_name 값 사용)
            if value is None:
                for canonical, aliases in self.FIELD_ALIASES.items():
                    if cell_key in aliases:
                        value = context.get(canonical)
                        if value is None:
                            for a in aliases:
                                value = context.get(a)
                                if value is not None:
                                    break
                        if value is not None:
                            break

            if value is None or value == "":
                skipped.append(f"{cell_key}({cell_addr})")
                continue

            # 날짜 처리
            if hasattr(value, "strftime"):
                value = value.strftime("%Y-%m-%d")

            # Decimal → int/float
            try:
                from decimal import Decimal
                if isinstance(value, Decimal):
                    value = int(value) if value == value.to_integral_value() else float(value)
            except Exception:
                pass

            try:
                anchor_addr = _anchor(ws, cell_addr)
                if _safe_set_cell(ws, anchor_addr, value):
                    written.append(f"{cell_key}={cell_addr}")
                else:
                    skipped.append(f"{cell_key}({cell_addr}→{anchor_addr}): label cell protected")
            except Exception as e:
                logger.warning("xlsx_cell_set_failed", cell=cell_addr, error=str(e))
                skipped.append(f"{cell_key}({cell_addr}): {e}")
