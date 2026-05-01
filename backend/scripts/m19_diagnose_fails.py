"""
M1.9 FAIL 진단 스크립트
- cell_year_2026, cell_quantity_5, cell_spec_kg, blacklist_clean FAIL 원인 분석
"""
import sys
sys.path.insert(0, "/app")

import openpyxl
from pathlib import Path

OUTPUT_DIR = Path("/app/outputs_verify/매핑검증_2026-04-30")

TARGETS = {
    "㈜동우": {
        "file_pattern": "㈜동우_quote_",
        "cells_to_check": ["A5"],          # issue_date
        "blacklist": [],
    },
    "지티정밀": {
        "file_pattern": "지티정밀_quote_",
        "cells_to_check": ["H3", "F11"],   # issue_date, spec
        "blacklist": [],
    },
    "신라정밀": {
        "file_pattern": "신라정밀_quote_",
        "cells_to_check": ["B5", "C5", "D5", "C12", "D12"],  # year/month/day, spec, quantity
        "blacklist": [],
    },
    "에스엠앤씨": {
        "file_pattern": "에스엠앤씨 _quote_",
        "cells_to_check": ["B1", "C1", "D1", "F9"],  # year/month/day, quantity
        "blacklist": [],
    },
    "선양에스피티": {
        "file_pattern": "선양에스피티_quote_",
        "cells_to_check": ["A1", "F9"],    # issue_date, quantity
        "blacklist": [],
    },
    "(주)옵토마린": {
        "file_pattern": "(주)옵토마린_quote_",
        "cells_to_check": ["B13", "C13", "D13", "J15"],  # year/month/day, quantity
        "blacklist": ["2017"],
    },
    "아이에이치켐": {
        "file_pattern": "아이에이치켐(IH CHEM) _quote_",
        "cells_to_check": ["AJ11"],  # quantity
        "blacklist": ["2022"],
    },
}

def read_specific_cells(xlsx_path: str, cells: list[str]) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    result = {}
    for cell_addr in cells:
        try:
            val = ws[cell_addr].value
            result[cell_addr] = repr(val)
        except Exception as e:
            result[cell_addr] = f"ERROR: {e}"
    wb.close()
    return result

def find_blacklist_in_file(xlsx_path: str, terms: list[str]) -> dict:
    """각 blacklist 단어가 어느 셀에 있는지 찾기"""
    if not terms:
        return {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    found = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_str = str(cell.value)
                    for term in terms:
                        if term in cell_str:
                            coord = f"{ws.title}!{cell.coordinate}"
                            found.setdefault(term, []).append(f"{coord}={repr(cell.value)}")
    wb.close()
    return found

def read_all_cells(xlsx_path: str) -> list[str]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    vals = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    vals.append(str(cell.value))
    wb.close()
    return vals

print("=" * 70)
print("M1.9 FAIL 진단 결과")
print("=" * 70)

for vendor_name, cfg in TARGETS.items():
    pattern = cfg["file_pattern"]
    files = sorted(OUTPUT_DIR.glob(f"{pattern}*.xlsx"))
    if not files:
        print(f"\n[{vendor_name}] ❌ 파일 없음 (패턴: {pattern})")
        continue

    f = files[0]  # 가장 첫 번째 파일만
    print(f"\n[{vendor_name}]  →  {f.name}")

    # 특정 셀 값 확인
    cell_vals = read_specific_cells(str(f), cfg["cells_to_check"])
    for addr, val in cell_vals.items():
        print(f"  셀 {addr:8s} = {val}")

    # 전체 셀 값에서 FAIL 키워드 확인
    all_vals = read_all_cells(str(f))
    has_2026 = any("2026" in v for v in all_vals)
    has_5 = "5" in all_vals or "5.0" in all_vals
    has_kg = any("kg" in v.lower() for v in all_vals)
    print(f"  cell_year_2026: {'PASS' if has_2026 else 'FAIL'} (전체 셀에 2026 {'있음' if has_2026 else '없음'})")
    print(f"  cell_quantity_5: {'PASS' if has_5 else 'FAIL'} (전체 셀에 '5' {'있음' if has_5 else '없음'})")
    print(f"  cell_spec_kg: {'PASS' if has_kg else 'FAIL'} (전체 셀에 'kg' {'있음' if has_kg else '없음'})")

    # blacklist 위치 찾기
    if cfg["blacklist"]:
        found = find_blacklist_in_file(str(f), cfg["blacklist"])
        for term, locations in found.items():
            print(f"  blacklist '{term}' 위치:")
            for loc in locations:
                print(f"    {loc}")
        if not found:
            print(f"  blacklist {cfg['blacklist']}: 없음")

print("\n" + "=" * 70)
