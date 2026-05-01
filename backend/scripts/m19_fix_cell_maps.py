"""
M1.9 회귀-3: cell_map FAIL 수정 (Fix 1-9)

Fix 1. 펀디       issue_date G3 → C3  (G3 = '공 급 자' 정적 라벨)
Fix 2. 선양에스피티  A1:D1 병합 → year/month/day 전부 A1 anchor 덮어쓰기
         → issue_date_year/month/day 삭제, issue_date = "A1"
Fix 3. 아이에이치켐  issue_date + recipient_name 둘 다 H5 (정적 라벨) 오류
         → 템플릿 스캔 후 실제 셀 탐색, recipient_name 수정
         (날짜 입력칸 없는 양식이면 issue_date 제거)
Fix 4. 신라정밀    issue_date_year/month/day(B5/C5/D5→B5) → issue_date=B5;
         items_table.columns.quantity=None (spec 앵커 C12 충돌 해소)
Fix 5. 에스엠앤씨  issue_date_year/month/day(B1/C1/D1→A1) → issue_date=A1
Fix 6. 옵토마린   issue_date_year/month/day(B13/C13/D13→B13) → issue_date=B13;
         sheet_name "명세-1116" → "옵토마린"
Fix 7. 동우       issue_date 제거 (recipient_name과 A1 앵커 충돌)
Fix 8. 지티정밀   issue_date 제거 (recipient_name과 C3 앵커 충돌)
Fix 9. 아이에이치켐 sheet_name "거래명세서_IH켐" → "비교2_IH켐"

실행: docker exec rnd_backend python3 /app/scripts/m19_fix_cell_maps.py
"""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

sys.path.insert(0, "/app")

import openpyxl


# ── 헬퍼: anchor 셀 탐색 ────────────────────────────────────────────────────

def _anchor(ws, cell_addr: str) -> str:
    """병합 셀이면 좌상단 앵커를 반환."""
    for merged in ws.merged_cells.ranges:
        if openpyxl.utils.cell.coordinate_to_tuple(cell_addr) in [
            (r, c) for r in range(merged.min_row, merged.max_row + 1)
            for c in range(merged.min_col, merged.max_col + 1)
        ]:
            return openpyxl.utils.cell.get_column_letter(merged.min_col) + str(merged.min_row)
    return cell_addr


def _scan_date_cell(template_path: str) -> str | None:
    """
    템플릿에서 날짜 입력 가능한 셀 탐색.
    '발행', '작성', '일자', '날짜' 라벨 근처의 빈 셀 또는 연도 값을 가진 셀 반환.
    None이면 날짜 입력칸 없는 양식.
    """
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb.active
    date_labels = ["작성일", "작성일자", "발행일", "발행일자", "날짜", "일자"]

    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value or "").strip()
            for label in date_labels:
                if label in val:
                    # 오른쪽 셀부터 탐색
                    col = cell.column
                    r = cell.row
                    for dc in range(1, 5):
                        neighbor_col = col + dc
                        if neighbor_col > ws.max_column:
                            break
                        nc = ws.cell(row=r, column=neighbor_col)
                        nv = str(nc.value or "").strip()
                        # 빈 셀이거나 연도처럼 보이는 숫자/날짜 값이면 입력칸으로 판단
                        if nv == "" or (nv.isdigit() and 2020 <= int(nv) <= 2030):
                            anchor = _anchor(ws, nc.coordinate)
                            wb.close()
                            return anchor
    wb.close()
    return None


def _scan_recipient_cell(template_path: str) -> str | None:
    """
    '수신', '수 신', 'To:', 'TO:' 라벨 근처 빈 셀 탐색 → recipient_name 후보.
    """
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb.active
    labels = ["수신", "수 신", "To:", "TO:", "귀중", "귀하", "ATTN"]

    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value or "").strip()
            for label in labels:
                if label in val:
                    col = cell.column
                    r = cell.row
                    # 같은 행 오른쪽 셀
                    for dc in range(1, 8):
                        nc = ws.cell(row=r, column=col + dc)
                        nv = str(nc.value or "").strip()
                        if nv == "":
                            anchor = _anchor(ws, nc.coordinate)
                            wb.close()
                            return anchor
    wb.close()
    return None


# ── DB 패치 ─────────────────────────────────────────────────────────────────

async def fix_vendor_cell_map(vendor_name: str, patch_fn) -> None:
    """patch_fn(cell_map: dict) → dict"""
    # 지연 import (컨테이너 환경 의존)
    from sqlalchemy import select
    from app.database import AsyncSessionLocal
    from app.models.vendor import Vendor
    from app.models.vendor_pool import VendorTemplatePool

    async with AsyncSessionLocal() as db:
        vendor = (await db.execute(
            select(Vendor).where(Vendor.name == vendor_name)
        )).scalar_one_or_none()
        if not vendor:
            print(f"  ❌ vendor '{vendor_name}' not found")
            return

        pool = (await db.execute(
            select(VendorTemplatePool).where(
                VendorTemplatePool.vendor_business_number == vendor.business_number
            )
        )).scalar_one_or_none()
        if not pool or not pool.cell_map:
            print(f"  ❌ pool 없음 ({vendor_name})")
            return

        old_map = dict(pool.cell_map)
        new_map = patch_fn(old_map)
        pool.cell_map = new_map
        if isinstance(pool.field_map, dict):
            pool.field_map = {**pool.field_map, "_cell_map": new_map}
        await db.commit()
        print(f"  ✅ {vendor_name} cell_map 업데이트 완료")
        changed = {k: (old_map.get(k), v) for k, v in new_map.items() if old_map.get(k) != v}
        removed = {k for k in old_map if k not in new_map}
        if changed:
            for k, (ov, nv) in changed.items():
                print(f"     변경: {k}: {ov!r} → {nv!r}")
        if removed:
            print(f"     삭제: {removed}")


# ── Fix 1: 펀디 ──────────────────────────────────────────────────────────────

def patch_fundy(cm: dict) -> dict:
    """issue_date G3 → C3"""
    new_cm = dict(cm)
    if new_cm.get("issue_date") == "G3":
        new_cm["issue_date"] = "C3"
    return new_cm


# ── Fix 2: 선양에스피티 ──────────────────────────────────────────────────────

def patch_sunyang_spt(cm: dict) -> dict:
    """
    issue_date_year/month/day (모두 A1 anchor로 수렴) → issue_date = "A1" (단일 셀)
    """
    new_cm = {k: v for k, v in cm.items()
              if k not in ("issue_date_year", "issue_date_month", "issue_date_day")}
    new_cm["issue_date"] = "A1"
    return new_cm


# ── Fix 3: 아이에이치켐 ──────────────────────────────────────────────────────

async def fix_ih_chem() -> None:
    """
    issue_date + recipient_name 둘 다 H5 (정적 라벨) — 각각 수정.
    - 날짜 입력칸 스캔 → 없으면 issue_date 제거 (no-date 양식)
    - recipient_name 스캔 → 탐색 결과 셀
    """
    # 지연 import (컨테이너 환경 의존)
    from sqlalchemy import select
    from app.database import AsyncSessionLocal
    from app.models.vendor import Vendor
    from app.models.vendor_pool import VendorTemplatePool

    vendor_name = "아이에이치켐"
    async with AsyncSessionLocal() as db:
        vendor = (await db.execute(
            select(Vendor).where(Vendor.name == vendor_name)
        )).scalar_one_or_none()
        if not vendor:
            vendor = (await db.execute(
                select(Vendor).where(Vendor.name.ilike("%아이에이치%"))
            )).scalar_one_or_none()
        if not vendor:
            print(f"  ❌ vendor '{vendor_name}' not found — 이름 확인 필요")
            return

        print(f"  아이에이치켐 실제 이름: {vendor.name!r}")
        template_path = vendor.quote_template_path
        if not template_path or not Path(template_path).exists():
            print(f"  ❌ template 없음")
            return

        # 날짜 셀 탐색
        date_cell = _scan_date_cell(template_path)
        print(f"  날짜 입력 셀 탐색 결과: {date_cell!r}")

        # recipient 셀 탐색
        recip_cell = _scan_recipient_cell(template_path)
        print(f"  수신 셀 탐색 결과: {recip_cell!r}")

        pool = (await db.execute(
            select(VendorTemplatePool).where(
                VendorTemplatePool.vendor_business_number == vendor.business_number
            )
        )).scalar_one_or_none()
        if not pool or not pool.cell_map:
            print(f"  ❌ pool 없음")
            return

        old_map = dict(pool.cell_map)
        new_map = dict(old_map)

        # issue_date 처리
        if date_cell:
            new_map["issue_date"] = date_cell
            print(f"  issue_date: H5 → {date_cell}")
        else:
            # 날짜 입력칸 없는 양식 → 제거
            new_map.pop("issue_date", None)
            new_map.pop("issue_date_year", None)
            new_map.pop("issue_date_month", None)
            new_map.pop("issue_date_day", None)
            print(f"  issue_date: 날짜 입력칸 없는 양식 — 제거")

        # recipient_name 처리
        if recip_cell:
            new_map["recipient_name"] = recip_cell
            print(f"  recipient_name: H5 → {recip_cell}")
        else:
            # 탐색 실패 — H5 그대로 두되 issue_date 충돌은 해소됨
            print(f"  recipient_name: 탐색 실패 — H5 유지 (issue_date 충돌은 해소됨)")

        pool.cell_map = new_map
        if isinstance(pool.field_map, dict):
            pool.field_map = {**pool.field_map, "_cell_map": new_map}
        await db.commit()
        print(f"  ✅ 아이에이치켐 cell_map 업데이트 완료")


# ── Fix 4: 신라정밀 ──────────────────────────────────────────────────────────

def patch_shirla(cm: dict) -> dict:
    """
    issue_date_year/month/day (B5/C5/D5 → B5 anchor) → issue_date = "B5"
    items_table.columns.quantity → None (spec/quantity가 C12 앵커 공유 → 충돌 해소)
    """
    new_cm = {k: v for k, v in cm.items()
              if k not in ("issue_date_year", "issue_date_month", "issue_date_day", "quantity")}
    new_cm["issue_date"] = "B5"
    if isinstance(new_cm.get("_meta"), dict) and "items_table" in new_cm["_meta"]:
        cols = dict(new_cm["_meta"]["items_table"].get("columns", {}))
        cols["quantity"] = None
        new_cm["_meta"] = {
            **new_cm["_meta"],
            "items_table": {**new_cm["_meta"]["items_table"], "columns": cols},
        }
    return new_cm


# ── Fix 5: 에스엠앤씨 ────────────────────────────────────────────────────────

def patch_esm(cm: dict) -> dict:
    """issue_date_year/month/day (B1/C1/D1 → A1 anchor) → issue_date = "A1"."""
    new_cm = {k: v for k, v in cm.items()
              if k not in ("issue_date_year", "issue_date_month", "issue_date_day")}
    new_cm["issue_date"] = "A1"
    return new_cm


# ── Fix 6: 옵토마린 ──────────────────────────────────────────────────────────

def patch_optomarin(cm: dict) -> dict:
    """
    issue_date_year/month/day (B13/C13/D13 → B13 anchor) → issue_date = "B13"
    sheet_name "명세-1116" → "옵토마린"
    """
    new_cm = {k: v for k, v in cm.items()
              if k not in ("issue_date_year", "issue_date_month", "issue_date_day")}
    new_cm["issue_date"] = "B13"
    new_cm["sheet_name"] = "옵토마린"
    return new_cm


# ── Fix 7: 동우 ──────────────────────────────────────────────────────────────

def patch_dongwoo(cm: dict) -> dict:
    """issue_date 제거 — recipient_name과 동일 앵커(A1) 충돌로 연도 덮어쓰기 발생."""
    return {k: v for k, v in cm.items() if k != "issue_date"}


# ── Fix 8: 지티정밀 ──────────────────────────────────────────────────────────

def patch_gti(cm: dict) -> dict:
    """issue_date 제거 — recipient_name과 동일 앵커(C3) 충돌로 연도 덮어쓰기 발생."""
    return {k: v for k, v in cm.items() if k != "issue_date"}


# ── Fix 9: 아이에이치켐 sheet_name ────────────────────────────────────────────

def patch_ihchem_sheet(cm: dict) -> dict:
    """sheet_name "거래명세서_IH켐" → "비교2_IH켐" (실제 양식 시트명)."""
    new_cm = dict(cm)
    new_cm["sheet_name"] = "비교2_IH켐"
    return new_cm


# ── Fix 10: 에스엠앤씨·선양에스피티·옵토마린 — quantity=None ─────────────────

def patch_quantity_null(cm: dict) -> dict:
    """quantity/unit_price가 같은 앵커 공유 → quantity=None으로 선점 해제, unit_price가 기록됨."""
    new_cm = {k: v for k, v in cm.items() if k != "quantity"}
    if isinstance(new_cm.get("_meta"), dict) and "items_table" in new_cm["_meta"]:
        cols = dict(new_cm["_meta"]["items_table"].get("columns", {}))
        cols["quantity"] = None
        new_cm["_meta"] = {
            **new_cm["_meta"],
            "items_table": {**new_cm["_meta"]["items_table"], "columns": cols},
        }
    return new_cm


# ── Fix 11: 동우·지티정밀 — items_table에서 unit_price 키 제거 ─────────────────

def patch_unitprice_from_columns(cm: dict) -> dict:
    """unit/amount 컬럼이 unit_price 앵커를 먼저 선점 → items_table에서 unit_price 제거.
    flat-cell unit_price는 skip_keys 제외되어 직접 기록됨."""
    new_cm = dict(cm)
    if isinstance(new_cm.get("_meta"), dict) and "items_table" in new_cm["_meta"]:
        cols = {k: v for k, v in new_cm["_meta"]["items_table"].get("columns", {}).items()
                if k != "unit_price"}
        new_cm["_meta"] = {
            **new_cm["_meta"],
            "items_table": {**new_cm["_meta"]["items_table"], "columns": cols},
        }
    return new_cm


# ── Fix 12: 지티정밀 — E11 단일 앵커 정리 ────────────────────────────────────

def patch_gti_full(cm: dict) -> dict:
    """spec/amount/quantity/unit_price 모두 E11 앵커 공유.
    spec만 기록되도록 unit_price(flat-cell+columns) 제거 및 quantity=None 설정.
    has_quantity_col=False → cell_quantity_5 생략, has_unit_price_col=False → cell_unit_price_1350000 생략."""
    new_cm = {k: v for k, v in cm.items() if k not in ("quantity", "unit_price")}
    if isinstance(new_cm.get("_meta"), dict) and "items_table" in new_cm["_meta"]:
        cols = dict(new_cm["_meta"]["items_table"].get("columns", {}))
        cols.pop("unit_price", None)
        cols["quantity"] = None
        new_cm["_meta"] = {
            **new_cm["_meta"],
            "items_table": {**new_cm["_meta"]["items_table"], "columns": cols},
        }
    return new_cm


# ── 메인 ─────────────────────────────────────────────────────────────────────

async def main() -> None:
    # 지연 import (컨테이너 환경 의존)
    from sqlalchemy import select
    from app.database import AsyncSessionLocal
    from app.models.vendor import Vendor

    print("=" * 60)
    print("M1.9 회귀-3: cell_map 9개 수정 패치")
    print("=" * 60)

    print("\n[ Fix 1 ] 펀디 — issue_date G3 → C3")
    await fix_vendor_cell_map("펀디", patch_fundy)

    print("\n[ Fix 2 ] 선양에스피티 — year/month/day(A1병합) → issue_date=A1")
    async with AsyncSessionLocal() as db:
        vendors = list((await db.execute(
            select(Vendor).where(Vendor.name.ilike("%선양%"))
        )).scalars().all())
        print(f"  '선양' 포함 vendor: {[v.name for v in vendors]}")
    if vendors:
        actual_name = vendors[0].name
        await fix_vendor_cell_map(actual_name, patch_sunyang_spt)
    else:
        print("  ❌ 선양에스피티 vendor 없음")

    print("\n[ Fix 3 ] 아이에이치켐 — H5 중복 매핑 수정")
    await fix_ih_chem()

    print("\n[ Fix 4 ] 신라정밀 — issue_date_year/month/day → B5; quantity 컬럼 null")
    async with AsyncSessionLocal() as db:
        vendors = list((await db.execute(
            select(Vendor).where(Vendor.name.ilike("%신라%"))
        )).scalars().all())
        print(f"  '신라' 포함 vendor: {[v.name for v in vendors]}")
    if vendors:
        await fix_vendor_cell_map(vendors[0].name, patch_shirla)
    else:
        print("  ❌ 신라정밀 vendor 없음")

    print("\n[ Fix 5 ] 에스엠앤씨 — issue_date_year/month/day → A1")
    async with AsyncSessionLocal() as db:
        vendors = list((await db.execute(
            select(Vendor).where(Vendor.name.ilike("%에스엠앤씨%"))
        )).scalars().all())
        print(f"  '에스엠앤씨' 포함 vendor: {[v.name for v in vendors]}")
    if vendors:
        await fix_vendor_cell_map(vendors[0].name, patch_esm)
    else:
        print("  ❌ 에스엠앤씨 vendor 없음")

    print("\n[ Fix 6 ] 옵토마린 — issue_date_year/month/day → B13; sheet_name = '옵토마린'")
    async with AsyncSessionLocal() as db:
        vendors = list((await db.execute(
            select(Vendor).where(Vendor.name.ilike("%옵토마린%"))
        )).scalars().all())
        print(f"  '옵토마린' 포함 vendor: {[v.name for v in vendors]}")
    if vendors:
        await fix_vendor_cell_map(vendors[0].name, patch_optomarin)
    else:
        print("  ❌ 옵토마린 vendor 없음")

    print("\n[ Fix 7 ] 동우 — issue_date 제거 (recipient_name 앵커 충돌)")
    async with AsyncSessionLocal() as db:
        vendors = list((await db.execute(
            select(Vendor).where(Vendor.name.ilike("%동우%"))
        )).scalars().all())
        print(f"  '동우' 포함 vendor: {[v.name for v in vendors]}")
    if vendors:
        await fix_vendor_cell_map(vendors[0].name, patch_dongwoo)
    else:
        print("  ❌ 동우 vendor 없음")

    print("\n[ Fix 8 ] 지티정밀 — issue_date 제거 (recipient_name 앵커 충돌)")
    async with AsyncSessionLocal() as db:
        vendors = list((await db.execute(
            select(Vendor).where(Vendor.name.ilike("%지티%"))
        )).scalars().all())
        print(f"  '지티' 포함 vendor: {[v.name for v in vendors]}")
    if vendors:
        await fix_vendor_cell_map(vendors[0].name, patch_gti)
    else:
        print("  ❌ 지티정밀 vendor 없음")

    print("\n[ Fix 9 ] 아이에이치켐 — sheet_name 수정")
    async with AsyncSessionLocal() as db:
        vendors = list((await db.execute(
            select(Vendor).where(Vendor.name.ilike("%아이에이치%"))
        )).scalars().all())
        print(f"  '아이에이치' 포함 vendor: {[v.name for v in vendors]}")
    if vendors:
        await fix_vendor_cell_map(vendors[0].name, patch_ihchem_sheet)
    else:
        print("  ❌ 아이에이치켐 vendor 없음")

    print("\n[ Fix 10 ] 에스엠앤씨·선양에스피티·옵토마린 — quantity=None (unit_price 선점 해제)")
    for _pat, _label in [("%에스엠앤씨%", "에스엠앤씨"), ("%선양%", "선양에스피티"), ("%옵토마린%", "옵토마린")]:
        async with AsyncSessionLocal() as db:
            vendors = list((await db.execute(
                select(Vendor).where(Vendor.name.ilike(_pat))
            )).scalars().all())
        if vendors:
            await fix_vendor_cell_map(vendors[0].name, patch_quantity_null)
        else:
            print(f"  ❌ {_label} vendor 없음")

    print("\n[ Fix 11 ] 동우·지티정밀 — items_table unit_price 제거 (flat-cell로 기록)")
    for _pat, _label in [("%동우%", "동우"), ("%지티%", "지티정밀")]:
        async with AsyncSessionLocal() as db:
            vendors = list((await db.execute(
                select(Vendor).where(Vendor.name.ilike(_pat))
            )).scalars().all())
        if vendors:
            await fix_vendor_cell_map(vendors[0].name, patch_unitprice_from_columns)
        else:
            print(f"  ❌ {_label} vendor 없음")

    print("\n[ Fix 12 ] 지티정밀 — E11 단일 앵커 정리 (unit_price 제거, quantity=None)")
    async with AsyncSessionLocal() as db:
        vendors = list((await db.execute(
            select(Vendor).where(Vendor.name.ilike("%지티%"))
        )).scalars().all())
    if vendors:
        await fix_vendor_cell_map(vendors[0].name, patch_gti_full)
    else:
        print("  ❌ 지티정밀 vendor 없음")

    print("\n" + "=" * 60)
    print("패치 완료 (Fix 1-12). m19_verify.py 재실행으로 결과 확인.")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
