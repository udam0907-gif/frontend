"""
병합 셀 구조 분석 — FAIL 원인 정확히 파악
각 FAIL vendor 템플릿의 병합 범위를 읽고,
cell_map 좌표 중 앵커가 아닌 셀 식별.
"""
import sys
sys.path.insert(0, "/app")
import asyncio
import json
from pathlib import Path
import openpyxl
from openpyxl.utils.cell import (
    coordinate_from_string, column_index_from_string, get_column_letter
)

TARGET_VENDORS = [
    "㈜동우", "지티정밀", "신라정밀", "에스엠앤씨 ", "선양에스피티",
    "(주)옵토마린", "아이에이치켐(IH CHEM) "
]

def get_anchor(ws, cell_addr: str) -> str:
    try:
        col_letter, row = coordinate_from_string(cell_addr)
        col_idx = column_index_from_string(col_letter)
    except Exception:
        return cell_addr
    for mr in ws.merged_cells.ranges:
        if (mr.min_row <= row <= mr.max_row
                and mr.min_col <= col_idx <= mr.max_col):
            return f"{get_column_letter(mr.min_col)}{mr.min_row}"
    return cell_addr

def analyze_template(template_path: str, cell_map: dict) -> dict:
    wb = openpyxl.load_workbook(template_path)
    sheet_name = cell_map.get("sheet_name")
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # 모든 병합 범위
    merged = [str(r) for r in ws.merged_cells.ranges]

    # cell_map 좌표별 앵커 확인
    key_anchors = {}
    skip = {"sheet_name", "_cell_map", "_mapping_status", "_meta"}
    for k, v in cell_map.items():
        if k in skip or not isinstance(v, str) or not v:
            continue
        anchor = get_anchor(ws, v)
        key_anchors[k] = {"cell": v, "anchor": anchor, "is_anchor": anchor == v}

    # 같은 앵커를 가리키는 키들 그룹화
    anchor_groups: dict[str, list[str]] = {}
    for k, info in key_anchors.items():
        anchor = info["anchor"]
        anchor_groups.setdefault(anchor, []).append(k)

    conflicts = {a: ks for a, ks in anchor_groups.items() if len(ks) > 1}

    wb.close()
    return {
        "sheet_name": ws.title,
        "merged_ranges": merged[:20],  # 최대 20개만
        "key_anchors": key_anchors,
        "anchor_conflicts": conflicts,  # 같은 앵커 → 여러 필드 충돌
    }

async def main():
    from sqlalchemy import select
    from app.database import AsyncSessionLocal
    from app.models.vendor import Vendor
    from app.models.vendor_pool import VendorTemplatePool

    results = {}
    async with AsyncSessionLocal() as db:
        vendors = list((await db.execute(select(Vendor))).scalars().all())
        pools = list((await db.execute(select(VendorTemplatePool))).scalars().all())

    pool_map = {p.vendor_business_number: p for p in pools}

    for v in vendors:
        if not any(t in v.name for t in ["동우", "지티정밀", "신라정밀", "에스엠앤씨", "선양에스피티", "옵토마린", "아이에이치"]):
            continue

        pool = pool_map.get(v.business_number)
        if not pool or not pool.cell_map:
            print(f"[{v.name}] cell_map 없음")
            continue

        tp = v.quote_template_path
        if not tp or not Path(tp).exists():
            print(f"[{v.name}] 템플릿 파일 없음: {tp}")
            continue

        print(f"\n{'='*60}")
        print(f"[{v.name}]  template: {Path(tp).name}")
        try:
            info = analyze_template(tp, pool.cell_map)
        except Exception as e:
            print(f"  ERROR: {e}")
            continue

        print(f"  시트: {info['sheet_name']}")
        print(f"  병합 범위 ({len(info['merged_ranges'])}개): {info['merged_ranges'][:10]}")

        print("\n  cell_map 좌표 → 앵커 매핑:")
        for k, i in info["key_anchors"].items():
            mark = "  " if i["is_anchor"] else "⚠️ "
            print(f"    {mark}{k}: {i['cell']} → anchor={i['anchor']}")

        if info["anchor_conflicts"]:
            print("\n  ❌ 앵커 충돌 (여러 필드 → 같은 셀):")
            for anchor, keys in info["anchor_conflicts"].items():
                print(f"    anchor={anchor} ← {keys}")
        else:
            print("\n  ✅ 앵커 충돌 없음")

        results[v.name] = info

    # JSON 저장
    out = Path("/app/outputs_verify/m19_merged_analysis.json")
    out.parent.mkdir(exist_ok=True)
    with open(str(out), "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\n\n분석 결과 저장: {out}")

asyncio.run(main())
