#!/usr/bin/env python3
"""
M1.9 검증 스크립트
==================
cell_map 영속화 + 옛 데이터 가드 + 출력 시점 API 재호출 0회 검증

실행 방법:
  docker exec rnd_backend python3 /app/scripts/m19_verify.py

출력 폴더:
  /app/outputs_verify/매핑검증_YYYY-MM-DD/
"""
from __future__ import annotations

import asyncio
import json
import os
import shutil
import sys
import uuid
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, "/app")
os.chdir("/app")

# ── 검증 입력값 ────────────────────────────────────────────────────────────────
TODAY = date.today().strftime("%Y-%m-%d")
OUTPUT_DIR = Path("/app/outputs_verify") / f"매핑검증_{TODAY}"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TEST_TITLE = "이산화티타늄(TiO2) 구매"
TEST_SPEC = "kg"
TEST_QUANTITY = 5
TEST_UNIT_PRICE = 1_350_000
TEST_AMOUNT = 6_750_000  # 5 × 1,350,000
TEST_DATE = "2026-04-01"
TEMP_BIZNUM = "999-99-99999"  # pool에 절대 없는 사업자번호 (MappingNotFoundError 테스트용)

# 실제 셀 값에서 확인할 패턴 (견적서 전용)
QUOTE_CELL_CHECKS = {
    "tio2_item_name": lambda vals: any(
        "tio2" in v.lower() or "tio" in v.lower() or "티타늄" in v.lower() or "이산화" in v.lower()
        for v in vals
    ),
    "quantity_5": lambda vals: "5" in vals or "5.0" in vals,
    "spec_kg": lambda vals: any("kg" in v.lower() for v in vals),
    "unit_price_1350000": lambda vals: any(
        "1350000" in v or "1,350,000" in v for v in vals
    ),
    "amount_6750000": lambda vals: any(
        "6750000" in v or "6,750,000" in v for v in vals
    ),
    "year_2026": lambda vals: any("2026" in v for v in vals),
}

# 옛 데이터 블랙리스트
BLACKLIST_TERMS = [
    "㈜씨엠", "한국섬유개발연구원", "이재웅",
    "PET Fiber", "POY 160/48",
    "1,2,3,4-Butanetetracarboxylic acid",
    "2016", "2017", "2018", "2022",
]

# transaction_statement 없음이 정상인 업체
TS_EXEMPT = {"민서정밀", "아이에이치캠"}

# ── 양식 정적 텍스트 보존 검증 (회귀-2 추가) ──────────────────────────────────
# vendor별 출력물에 반드시 존재해야 하는 정적 라벨·푸터 키워드.
# cell_map 초기화 가드가 정적 셀을 지우면 이 키워드들이 사라짐.
VENDOR_STATIC_KEYWORDS: dict[str, list[str]] = {
    "민서정밀": [
        "504-31-43112",      # 사업자등록번호 (vendor 자기 정보)
        "공장도착도 공급가임",  # 정적 푸터
        "오늘도 좋은 하루",   # 정적 인사말
        "등록번호",            # 정적 라벨
    ],
    "㈜대신테크젠": [
        "515-81-47073",      # 사업자등록번호
        "감사합니다",          # 정적 푸터
        "귀사의 성공",         # 정적 인사말
    ],
    "펀디": [
        "227-31-06281",      # 사업자등록번호
        "공 급 자",           # 정적 라벨 (양식 레이아웃)
    ],
    "태산물산": [
        "524-27-00949",      # 사업자등록번호
        "농협:351-1144-7422-93",  # 정적 입금정보
    ],
}

# 시트명 교차오염 검증 예외: 해당 vendor의 템플릿 시트명에 타 업체명이 포함될 수 있는 알려진 케이스.
# 에스와이케미칼 TS 템플릿 시트명 = "경구산업 거래명세표(케이테크) (2)" — 발주처 이름이 포함된 레거시 파일명.
SHEET_NAME_EXCEPTIONS: dict[str, list[str]] = {
    "에스와이케미칼": ["경구산업"],
}

# vendor 자기 정보 보존 검증: (사업자번호, 셀에 실제 존재하는 vendor 식별 문자열)
# 회사명이 이미지에만 있는 경우 대표자명 등 셀에 실제 존재하는 값으로 대체
VENDOR_INFO_CHECKS: dict[str, tuple[str, str]] = {
    "민서정밀":    ("504-31-43112", "민서정밀"),
    "㈜대신테크젠": ("515-81-47073", "황인성"),    # 회사명은 이미지 전용, 대표자명으로 대체
    "펀디":        ("227-31-06281", "펀  디"),     # 셀에 이중 공백 포함
    "태산물산":    ("524-27-00949", "태산물산"),
}


# ── 회귀-3: 우리 회사 영역 / 라인아이템 / 합계 검증 ──────────────────────────────
# 회사 설정에서 로드한 우리 회사 정보가 견적서 출력물에 반영됐는지 확인.
# (스크립트 실행 시 DB에서 동적으로 로드)
# recipient 영역 검증은 quote 문서에만 적용.

def check_recipient_area(
    cell_values: list[str],
    company_biznum: str,
    company_name: str,
    cell_map: dict,
) -> list[tuple[str, bool, str]]:
    """우리 회사 사업자번호·상호가 출력물 셀에 존재하는지 확인.
    - company_name(귀하) 검사: 항상 수행 (recipient_name 셀에 회사명 포함)
    - biznum 검사: cell_map에 recipient_business_number가 있을 때만 수행
    """
    all_text = " ".join(cell_values)
    results = []
    # 사업자번호: cell_map이 recipient_business_number 셀을 인식한 경우만 검사
    if company_biznum and cell_map.get("recipient_business_number"):
        found = company_biznum.replace("-", "") in all_text.replace("-", "")
        results.append((
            "recipient_biznum_filled",
            found,
            "" if found else f"우리 회사 사업자번호 {company_biznum!r} 미입력",
        ))
    # 회사명: recipient_name 셀에 회사명이 들어가므로 항상 검사
    if company_name:
        found = company_name in all_text
        results.append((
            "recipient_company_filled",
            found,
            "" if found else f"우리 회사명 {company_name!r} 미입력 (recipient_name 미설정 확인)",
        ))
    return results


def check_lineitem_count(cell_values: list[str], expected_count: int) -> tuple[str, bool, str]:
    """입력한 라인아이템 수만큼 품목명이 셀에 존재하는지 확인 (최소 기준)."""
    item_hits = sum(
        1 for v in cell_values
        if "tio2" in v.lower() or "티타늄" in v.lower() or "이산화" in v.lower()
    )
    passed = item_hits >= expected_count
    return (
        "lineitem_count_ge_expected",
        passed,
        "" if passed else f"품목 셀 {item_hits}개 < 기대 {expected_count}개",
    )


def check_total_amount(cell_values: list[str]) -> tuple[str, bool, str]:
    """total_amount(6,750,000)이 출력물에 존재하는지 확인."""
    all_text = " ".join(cell_values)
    found = "6750000" in all_text or "6,750,000" in all_text
    return (
        "total_amount_filled",
        found,
        "" if found else "총합계 6,750,000 셀에 미입력",
    )


# ── 결과 수집 ──────────────────────────────────────────────────────────────────

class Results:
    def __init__(self) -> None:
        self._cases: list[dict] = []
        self.analyze_calls: list[str] = []

    def add(self, vendor: str, doc_type: str, check: str, passed: bool, detail: str = "") -> None:
        icon = "✅" if passed else "❌"
        print(f"      {icon} [{doc_type}] {check}" + (f" → {detail}" if detail else ""))
        self._cases.append(
            {"vendor": vendor, "doc_type": doc_type, "check": check, "passed": passed, "detail": detail}
        )

    @property
    def total(self) -> int:
        return len(self._cases)

    @property
    def n_pass(self) -> int:
        return sum(1 for c in self._cases if c["passed"])

    @property
    def n_fail(self) -> int:
        return self.total - self.n_pass

    @property
    def overall_pass(self) -> bool:
        return self.n_fail == 0 and len(self.analyze_calls) == 0

    def save_report(self) -> Path:
        report_path = OUTPUT_DIR / "m19_report.json"
        with open(str(report_path), "w", encoding="utf-8") as f:
            json.dump(
                {
                    "timestamp": datetime.now().isoformat(),
                    "analyze_calls_during_output": len(self.analyze_calls),
                    "analyze_call_paths": self.analyze_calls,
                    "total": self.total,
                    "passed": self.n_pass,
                    "failed": self.n_fail,
                    "overall": "PASS" if self.overall_pass else "FAIL",
                    "cases": self._cases,
                },
                f,
                ensure_ascii=False,
                indent=2,
            )
        return report_path

    def print_summary(self) -> None:
        sep = "=" * 70
        print(f"\n{sep}")
        print("M1.9 최종 검증 결과")
        print(sep)
        fails = [c for c in self._cases if not c["passed"]]
        if fails:
            print("\n❌ 실패 케이스:")
            for c in fails:
                print(f"  [{c['vendor']} / {c['doc_type']}] {c['check']}")
                if c["detail"]:
                    print(f"    → {c['detail']}")
        print(f"\n총 {self.total}건  통과 {self.n_pass}건  실패 {self.n_fail}건")
        print(f"출력 시점 analyze() 호출: {len(self.analyze_calls)}회")
        if self.analyze_calls:
            for p in self.analyze_calls:
                print(f"  - {p}")
        verdict = "✅ PASS" if self.overall_pass else "❌ FAIL"
        print(f"\n종합 판정: {verdict}")
        print(sep)


# ── 셀 읽기 헬퍼 ───────────────────────────────────────────────────────────────

def read_xlsx_cell_values(xlsx_path: str) -> tuple[list[str], list[str]]:
    """openpyxl로 모든 셀 값을 읽어 (str_values, sheet_names) 반환."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet_names = list(wb.sheetnames)
    values: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values.append(str(cell.value))
    wb.close()
    return values, sheet_names


# ── 메인 ──────────────────────────────────────────────────────────────────────

async def main() -> bool:
    # 지연 import (컨테이너 환경 의존)
    from sqlalchemy import select
    from app.database import AsyncSessionLocal
    from app.models.vendor import Vendor
    from app.models.vendor_pool import VendorTemplatePool
    from app.models.project import Project
    from app.models.expense import ExpenseItem
    from app.models.company_setting import CompanySetting
    from app.models.enums import CategoryType
    from app.services.document_set_service import DocumentSetService
    from app.services.xlsx_cell_mapper import XlsxCellMapper
    from app.services.llm_service import get_llm_service
    from app.config import settings

    res = Results()
    print(f"\n{'='*70}")
    print(f"M1.9 검증 시작: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"출력 폴더: {OUTPUT_DIR}")
    print(f"{'='*70}")

    # ── Phase 0: Preflight ────────────────────────────────────────────────────
    print("\n[ Phase 0 ] Preflight")

    vendor_count = 0
    vendor_rows: list[Vendor] = []
    project = None
    our_company_name = ""

    async with AsyncSessionLocal() as db:
        # magic number 금지 — vendors 테이블에서 동적 조회
        vendor_rows = list(
            (await db.execute(
                select(Vendor)
                .order_by(Vendor.created_at)
            )).scalars().all()
        )
        vendor_count = len(vendor_rows)
        print(f"  Vendor 수 (DB 동적, 전체): {vendor_count}개")
        for v in vendor_rows:
            print(f"    · {v.name:20s} biz={v.business_number} quote={'O' if v.quote_template_path else 'X'} ts={'O' if v.transaction_statement_path else 'X'}")

        pool_rows = list((await db.execute(select(VendorTemplatePool))).scalars().all())
        print(f"\n  vendor_template_pool: {len(pool_rows)}행")
        for p in pool_rows:
            print(f"    · {p.vendor_name:20s} cell_map={'있음' if p.cell_map else '없음'}")

        project = (await db.execute(select(Project).limit(1))).scalar_one_or_none()
        print(f"\n  project: {project.name if project else '없음'}")

        cs = (await db.execute(
            select(CompanySetting).where(CompanySetting.company_id == "default")
        )).scalar_one_or_none()
        our_company_name = cs.company_name if cs else ""
        our_company_biznum = cs.company_registration_number if cs else ""
        print(f"  우리 회사명: {our_company_name!r}  사업자번호: {our_company_biznum!r}")

    if not vendor_rows:
        print("  ❌ Global vendor 없음 — 검증 불가")
        res.print_summary()
        return False

    # ── Phase 1: cell_map 사전 등록 (등록 시점 analyze() — 정당) ─────────────
    print("\n[ Phase 1 ] cell_map 사전 등록 (등록 시점 analyze())")

    mapper = XlsxCellMapper(get_llm_service())

    async with AsyncSessionLocal() as db:
        for vendor in vendor_rows:
            fp = vendor.quote_template_path
            if not fp or not Path(fp).exists():
                print(f"  {vendor.name}: quote_template 없음 — skip")
                continue
            ext = Path(fp).suffix.lower()
            if ext not in (".xlsx", ".xls"):
                print(f"  {vendor.name}: {ext} 형식 — XLSX 아님, skip")
                continue

            pool = (await db.execute(
                select(VendorTemplatePool).where(
                    VendorTemplatePool.vendor_business_number == vendor.business_number
                )
            )).scalar_one_or_none()

            if pool and pool.cell_map:
                print(f"  {vendor.name}: cell_map 재사용 ({len(pool.cell_map)}키)")
                continue

            print(f"  {vendor.name}: analyze() 호출 (등록 시점)")
            try:
                remap = await mapper.analyze(fp)
                cell_map = remap.get("cell_map", {})
                if not cell_map:
                    print(f"    ⚠️  cell_map 비어있음")
                    continue
                if pool:
                    pool.cell_map = cell_map
                    pool.field_map = {**pool.field_map, "_cell_map": cell_map, "_mapping_status": "auto_mapped"}
                else:
                    pool = VendorTemplatePool(
                        id=uuid.uuid4(),
                        vendor_business_number=vendor.business_number,
                        vendor_name=vendor.name,
                        file_format=ext.lstrip("."),
                        layout_map={},
                        render_profile={},
                        field_map={"_cell_map": cell_map, "_mapping_status": "auto_mapped"},
                        cell_map=cell_map,
                        sample_file_path=fp,
                    )
                    db.add(pool)
                await db.flush()
                print(f"    → cell_map 저장 ({len(cell_map)}키)")
            except Exception as e:
                print(f"    ⚠️  analyze() 실패: {e}")
        await db.commit()

    # ── Phase 2: analyze() 호출 카운터 설치 (출력 시점 감시) ─────────────────
    print("\n[ Phase 2 ] analyze() 카운터 설치 — 출력 시점 호출 감시")
    _original_analyze = XlsxCellMapper.analyze

    async def _counting_analyze(self_inner, file_path: str, *a, **kw):
        res.analyze_calls.append(file_path)
        return await _original_analyze(self_inner, file_path, *a, **kw)

    XlsxCellMapper.analyze = _counting_analyze
    print("  monkeypatched ✅")

    # ── Phase 3: vendor별 문서세트 생성 + 검증 ───────────────────────────────
    print("\n[ Phase 3 ] vendor별 문서세트 생성 + 셀 검증")

    svc = DocumentSetService(settings.storage_documents_path)
    created_expense_ids: list[uuid.UUID] = []

    # XLSX quote_template이 있는 vendor 목록 (동적)
    xlsx_vendors = [
        v for v in vendor_rows
        if v.quote_template_path
        and Path(v.quote_template_path).suffix.lower() in (".xlsx", ".xls")
        and Path(v.quote_template_path).exists()
    ]

    for main_vendor in xlsx_vendors:
        compare_vendor = next((v for v in xlsx_vendors if v.id != main_vendor.id), None)
        print(f"\n  ── {main_vendor.name} (비교: {compare_vendor.name if compare_vendor else 'None'}) ──")

        if not project:
            print("  ⚠️  project 없음 — skip")
            continue

        # temp expense 생성
        async with AsyncSessionLocal() as db:
            expense = ExpenseItem(
                id=uuid.uuid4(),
                project_id=project.id,
                category_type=CategoryType.materials,
                title=TEST_TITLE,
                amount=Decimal(str(TEST_AMOUNT)),
                expense_date=TEST_DATE,
                vendor_name=main_vendor.name,
                vendor_registration_number=main_vendor.business_number,
                metadata_={
                    "vendor_id": str(main_vendor.id),
                    "compare_vendor_id": str(compare_vendor.id) if compare_vendor else None,
                    "line_items": [
                        {
                            "item_name": TEST_TITLE,
                            "spec": TEST_SPEC,
                            "quantity": TEST_QUANTITY,
                            "unit_price": TEST_UNIT_PRICE,
                            "amount": TEST_AMOUNT,
                            "remark": "",
                        }
                    ],
                },
            )
            db.add(expense)
            await db.flush()
            expense_id = expense.id
            await db.commit()
        created_expense_ids.append(expense_id)

        # generate_set()
        async with AsyncSessionLocal() as db:
            try:
                gen_result = await svc.generate_set(expense_id, db)
                await db.commit()
            except Exception as e:
                print(f"  ❌ generate_set() 예외: {e}")
                res.add(main_vendor.name, "ALL", "generate_set_no_exception", False, str(e))
                continue

        # vendor cell_map 조회 (검증 조건부 판단용)
        vendor_cell_map: dict = {}
        async with AsyncSessionLocal() as db:
            pool_row = (await db.execute(
                select(VendorTemplatePool).where(
                    VendorTemplatePool.vendor_business_number == main_vendor.business_number
                )
            )).scalar_one_or_none()
            if pool_row and pool_row.cell_map:
                vendor_cell_map = pool_row.cell_map

        # 각 문서 검증
        for item in gen_result.items:
            doc_type = item.document_type.value

            # TS exempt 처리
            if doc_type == "transaction_statement" and main_vendor.name in TS_EXEMPT:
                if item.status in ("vendor_file_missing", "mapping_required"):
                    res.add(main_vendor.name, doc_type, "ts_exempt_skip", True, "미등록 정상 생략")
                    continue

            # mapping_required
            if item.status == "mapping_required":
                res.add(main_vendor.name, doc_type, "cell_map_available", False, item.error_message or "")
                continue

            # vendor_file_missing
            if item.status == "vendor_file_missing":
                expected = (
                    doc_type == "transaction_statement" and main_vendor.name in TS_EXEMPT
                )
                res.add(main_vendor.name, doc_type, "file_present_or_exempt", expected, item.error_message or "")
                continue

            # XLSX 출력 파일 검증
            if doc_type in ("quote", "comparative_quote", "transaction_statement"):
                if not item.output_path or not Path(item.output_path).exists():
                    res.add(main_vendor.name, doc_type, "output_file_exists", False, str(item.output_path))
                    continue

                # 복사본 저장
                dest = OUTPUT_DIR / f"{main_vendor.name}_{doc_type}_{Path(item.output_path).name}"
                shutil.copy2(item.output_path, str(dest))

                # 비-XLSX 파일(DOCX 등)은 셀 검증 건너뜀
                out_ext = Path(item.output_path).suffix.lower()
                if out_ext not in (".xlsx", ".xls", ".xlsm", ".xltx"):
                    res.add(main_vendor.name, doc_type, "non_xlsx_passthrough", True,
                            f"DOCX/PDF 등 비-XLSX 출력 ({out_ext}) — 셀 검증 생략")
                    continue

                try:
                    cell_values, sheet_names = read_xlsx_cell_values(item.output_path)
                except Exception as e:
                    res.add(main_vendor.name, doc_type, "xlsx_readable", False, str(e))
                    continue

                res.add(main_vendor.name, doc_type, "xlsx_readable", True)

                # 키워드 검증 (견적서만)
                if doc_type == "quote":
                    has_quantity_col = bool(
                        vendor_cell_map.get("quantity") or
                        (isinstance(vendor_cell_map.get("_meta"), dict) and
                         vendor_cell_map["_meta"].get("items_table", {}).get("columns", {}).get("quantity"))
                    )
                    # 날짜 입력칸 보유 여부: cell_map에 issue_date 또는 year/month/day 중 하나라도 있으면 True
                    has_date_field = bool(
                        vendor_cell_map.get("issue_date") or
                        vendor_cell_map.get("issue_date_year") or
                        vendor_cell_map.get("issue_date_month") or
                        vendor_cell_map.get("issue_date_day")
                    )
                    for check_name, check_fn in QUOTE_CELL_CHECKS.items():
                        # quantity_5 검사: 템플릿에 quantity 컬럼이 없으면 생략
                        if check_name == "quantity_5" and not has_quantity_col:
                            res.add(main_vendor.name, doc_type, f"cell_{check_name}",
                                    True, "quantity 컬럼 없는 양식 — 검사 생략")
                            continue
                        # year_2026 검사: 날짜 입력칸 없는 양식이면 생략
                        if check_name == "year_2026" and not has_date_field:
                            res.add(main_vendor.name, doc_type, f"cell_{check_name}",
                                    True, "날짜 입력칸 없는 양식 — 검사 생략")
                            continue
                        passed = check_fn(cell_values)
                        res.add(main_vendor.name, doc_type, f"cell_{check_name}", passed)

                    # 양식 정적 텍스트 보존 검증 (회귀-2 추가)
                    static_kws = VENDOR_STATIC_KEYWORDS.get(main_vendor.name, [])
                    if static_kws:
                        all_text = " ".join(cell_values)
                        for kw in static_kws:
                            found = kw in all_text
                            res.add(
                                main_vendor.name, doc_type,
                                f"static_preserved:{kw[:20]}",
                                found,
                                "" if found else f"양식 정적 텍스트 소실: {kw!r}",
                            )
                    # else: 키워드 미정의 vendor는 검사 생략 (PASS — 회귀-3 신규 vendor)

                    # vendor 자기 정보 보존 검증 (회귀-2 추가)
                    info_check = VENDOR_INFO_CHECKS.get(main_vendor.name)
                    if info_check:
                        biz_no, name_part = info_check
                        all_text = " ".join(cell_values)
                        res.add(
                            main_vendor.name, doc_type, "vendor_biznum_preserved",
                            biz_no in all_text,
                            "" if biz_no in all_text else f"사업자번호 {biz_no!r} 소실",
                        )
                        res.add(
                            main_vendor.name, doc_type, "vendor_name_preserved",
                            name_part in all_text,
                            "" if name_part in all_text else f"회사명 {name_part!r} 소실",
                        )

                    # 우리 회사(공급받는자) 영역 검증 (회귀-3 추가)
                    if our_company_name or our_company_biznum:
                        for chk_name, chk_pass, chk_detail in check_recipient_area(
                            cell_values, our_company_biznum, our_company_name, vendor_cell_map
                        ):
                            res.add(main_vendor.name, doc_type, chk_name, chk_pass, chk_detail)
                    else:
                        res.add(main_vendor.name, doc_type, "recipient_area_skip",
                                True, "회사 설정 미입력 — 검증 생략")

                    # 라인아이템 개수 검증 (회귀-3 추가)
                    chk_name, chk_pass, chk_detail = check_lineitem_count(cell_values, 1)
                    res.add(main_vendor.name, doc_type, chk_name, chk_pass, chk_detail)

                    # 합계 금액 검증 (회귀-3 추가)
                    chk_name, chk_pass, chk_detail = check_total_amount(cell_values)
                    res.add(main_vendor.name, doc_type, chk_name, chk_pass, chk_detail)

                # 블랙리스트 검증 (모든 XLSX 문서)
                all_text = " ".join(cell_values)
                hits = [t for t in BLACKLIST_TERMS if t in all_text]
                res.add(
                    main_vendor.name, doc_type, "blacklist_clean",
                    len(hits) == 0,
                    f"발견: {hits}" if hits else "",
                )

                # 시트명 교차 오염 검증
                # comparative_quote는 compare_vendor 템플릿 사용 → 첫 시트를 compare_vendor.name으로
                # rename하므로 compare_vendor 이름은 허용
                # SHEET_NAME_EXCEPTIONS: 템플릿 파일명에 발주처 이름이 박혀있는 레거시 케이스 허용
                wrong = []
                sn_exceptions = SHEET_NAME_EXCEPTIONS.get(main_vendor.name, [])
                for other in vendor_rows:
                    if other.id == main_vendor.id:
                        continue
                    if doc_type == "comparative_quote" and compare_vendor and other.id == compare_vendor.id:
                        continue
                    for sn in sheet_names:
                        # 예외 목록에 있는 vendor 이름이 시트명에 포함된 경우 무시
                        if any(exc in sn for exc in sn_exceptions):
                            continue
                        if other.name in sn:
                            wrong.append(f"시트'{sn}'에 '{other.name}' 포함")
                res.add(
                    main_vendor.name, doc_type, "sheet_name_no_cross_contamination",
                    len(wrong) == 0,
                    "; ".join(wrong) if wrong else "",
                )

            else:
                # passthrough 문서
                if item.output_path and Path(item.output_path).exists():
                    res.add(main_vendor.name, doc_type, "passthrough_exists", True)
                elif item.status != "vendor_file_missing":
                    res.add(
                        main_vendor.name, doc_type, "output_or_status_ok",
                        False, f"status={item.status}",
                    )

    # ── Phase 4: MappingNotFoundError 검증 ───────────────────────────────────
    print("\n[ Phase 4 ] MappingNotFoundError 검증 (pool 미등록 업체)")

    if project and xlsx_vendors:
        temp_vendor_id = uuid.uuid4()
        temp_expense_id = uuid.uuid4()

        async with AsyncSessionLocal() as db:
            temp_v = Vendor(
                id=temp_vendor_id,
                project_id=None,
                name="_M19_TEMP_VENDOR_",
                vendor_category="매입처",
                business_number=TEMP_BIZNUM,
                quote_template_path=xlsx_vendors[0].quote_template_path,
            )
            db.add(temp_v)
            temp_e = ExpenseItem(
                id=temp_expense_id,
                project_id=project.id,
                category_type=CategoryType.materials,
                title="매핑 미등록 테스트",
                amount=Decimal("1000000"),
                expense_date=TEST_DATE,
                vendor_name=temp_v.name,
                metadata_={
                    "vendor_id": str(temp_vendor_id),
                    "compare_vendor_id": None,
                    "line_items": [{"item_name": "test", "quantity": 1, "unit_price": 1000000, "amount": 1000000}],
                },
            )
            db.add(temp_e)
            await db.commit()

        created_expense_ids.append(temp_expense_id)

        async with AsyncSessionLocal() as db:
            gen_r = await svc.generate_set(temp_expense_id, db)
            await db.commit()

        quote_item = next(
            (i for i in gen_r.items if i.document_type.value == "quote"), None
        )
        expected_status = "mapping_required"
        actual_status = quote_item.status if quote_item else "not_found"
        res.add(
            "_M19_TEMP_", "quote", "mapping_not_found_returns_mapping_required",
            actual_status == expected_status,
            f"실제={actual_status}",
        )

        # temp vendor 정리 (expense는 Phase 6에서 일괄 삭제)
        async with AsyncSessionLocal() as db:
            tv = (await db.execute(
                select(Vendor).where(Vendor.business_number == TEMP_BIZNUM)
            )).scalar_one_or_none()
            if tv:
                await db.delete(tv)
            await db.commit()
    else:
        print("  ⚠️  project 없거나 XLSX vendor 없음 — skip")

    # ── Phase 5: analyze() 카운터 복원 + 검증 ────────────────────────────────
    print("\n[ Phase 5 ] analyze() 호출 횟수 (출력 시점)")
    XlsxCellMapper.analyze = _original_analyze

    if res.analyze_calls:
        print(f"  ❌ {len(res.analyze_calls)}회 호출 감지:")
        for p in res.analyze_calls:
            print(f"    - {p}")
    else:
        print("  ✅ 0회 — 정책 준수 (등록 시점 외 API 호출 없음)")

    # ── Phase 6: 임시 데이터 정리 ─────────────────────────────────────────────
    print("\n[ Phase 6 ] 임시 expense_item 정리")
    async with AsyncSessionLocal() as db:
        for eid in created_expense_ids:
            e = (await db.execute(
                select(ExpenseItem).where(ExpenseItem.id == eid)
            )).scalar_one_or_none()
            if e:
                await db.delete(e)
        await db.commit()
    print(f"  {len(created_expense_ids)}개 삭제 완료")

    # ── 최종 보고 ─────────────────────────────────────────────────────────────
    res.print_summary()
    report_path = res.save_report()
    print(f"\n보고서: {report_path}")

    return res.overall_pass


if __name__ == "__main__":
    success = asyncio.run(main())
    sys.exit(0 if success else 1)
