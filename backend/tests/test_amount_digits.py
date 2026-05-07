"""v6 _fill_amount_digits — γ 정책 단위 테스트.

γ 정책: 가장 큰 유효 자리(first_nonzero) 이전은 빈칸(None),
이후는 0 포함 모두 정수로 박기. 0원은 마지막 자리에만 0 표시.
"""

from __future__ import annotations

from openpyxl import Workbook

from app.services.xlsx_document_filler import _fill_amount_digits


class TestFillAmountDigits:
    def test_275000_to_9_digits(self) -> None:
        """275,000원 (γ): E/F/G 빈칸, H/I/J/K/L/M = 2/7/5/0/0/0."""
        wb = Workbook()
        ws = wb.active
        filled = _fill_amount_digits(ws, 275000, "E13")
        # "000275000" — first_nonzero=index 3 (H, 십만자리)
        assert ws["E13"].value is None  # 억자리 — 빈칸
        assert ws["F13"].value is None  # 천만자리
        assert ws["G13"].value is None  # 백만자리
        assert ws["H13"].value == 2
        assert ws["I13"].value == 7
        assert ws["J13"].value == 5
        assert ws["K13"].value == 0      # 백자리 — 유효자리 이후라 0 박힘
        assert ws["L13"].value == 0      # 십자리
        assert ws["M13"].value == 0      # 일자리
        assert filled == ["E13", "F13", "G13", "H13", "I13", "J13", "K13", "L13", "M13"]

    def test_100000000_to_9_digits(self) -> None:
        """1억 (γ): first_nonzero=0이라 9자리 모두 박힘 (1,0,0,0,0,0,0,0,0)."""
        wb = Workbook()
        ws = wb.active
        _fill_amount_digits(ws, 100000000, "E13")
        assert ws["E13"].value == 1
        for col in "FGHIJKLM":
            assert ws[f"{col}13"].value == 0

    def test_5_won_to_9_digits(self) -> None:
        """5원 (γ): first_nonzero=8(M) → M=5, 나머지 빈칸."""
        wb = Workbook()
        ws = wb.active
        _fill_amount_digits(ws, 5, "E13")
        assert ws["M13"].value == 5
        for col in "EFGHIJKL":
            assert ws[f"{col}13"].value is None

    def test_zero_amount(self) -> None:
        """0원 (γ fallback): M자리에 0, 나머지 빈칸."""
        wb = Workbook()
        ws = wb.active
        filled = _fill_amount_digits(ws, 0, "E13")
        for col in "EFGHIJKL":
            assert ws[f"{col}13"].value is None
        assert ws["M13"].value == 0
        assert len(filled) == 9

    def test_none_returns_empty(self) -> None:
        wb = Workbook()
        ws = wb.active
        filled = _fill_amount_digits(ws, None, "E13")
        assert filled == []

    def test_negative_returns_empty(self) -> None:
        wb = Workbook()
        ws = wb.active
        filled = _fill_amount_digits(ws, -100, "E13")
        assert filled == []

    def test_overflow_truncates_to_lower_digits(self) -> None:
        """9자리 초과 → 하위 9자리만. γ: first_nonzero=0(E)이라 모두 박힘."""
        wb = Workbook()
        ws = wb.active
        _fill_amount_digits(ws, 12_345_678_901, "E13")  # 11자리 → 하위 9자리 "345678901"
        assert ws["E13"].value == 3
        assert ws["F13"].value == 4
        assert ws["G13"].value == 5
        assert ws["H13"].value == 6
        assert ws["I13"].value == 7
        assert ws["J13"].value == 8
        assert ws["K13"].value == 9
        assert ws["L13"].value == 0   # γ: 유효자리 이후의 0은 박힘
        assert ws["M13"].value == 1

    def test_full_9_digits(self) -> None:
        """딱 9자리 (123,456,789)."""
        wb = Workbook()
        ws = wb.active
        _fill_amount_digits(ws, 123456789, "E13")
        assert ws["E13"].value == 1
        assert ws["F13"].value == 2
        assert ws["G13"].value == 3
        assert ws["H13"].value == 4
        assert ws["I13"].value == 5
        assert ws["J13"].value == 6
        assert ws["K13"].value == 7
        assert ws["L13"].value == 8
        assert ws["M13"].value == 9

    def test_different_start_cell(self) -> None:
        """시작 셀이 E13가 아닌 다른 좌표여도 정상."""
        wb = Workbook()
        ws = wb.active
        _fill_amount_digits(ws, 12345, "B5")
        # 9자리 zero-pad: "000012345"
        # B5=0 C5=0 D5=0 E5=0 F5=1 G5=2 H5=3 I5=4 J5=5
        assert ws["F5"].value == 1
        assert ws["G5"].value == 2
        assert ws["H5"].value == 3
        assert ws["I5"].value == 4
        assert ws["J5"].value == 5
        for col in "BCDE":
            assert ws[f"{col}5"].value is None

    def test_invalid_start_cell(self) -> None:
        """좌표 파싱 실패 시 빈 리스트."""
        wb = Workbook()
        ws = wb.active
        assert _fill_amount_digits(ws, 1000, "") == []
        assert _fill_amount_digits(ws, 1000, "E") == []
        assert _fill_amount_digits(ws, 1000, "13") == []

    def test_custom_num_digits(self) -> None:
        """num_digits=5 → 5자리 박기."""
        wb = Workbook()
        ws = wb.active
        _fill_amount_digits(ws, 12345, "A1", num_digits=5)
        assert ws["A1"].value == 1
        assert ws["B1"].value == 2
        assert ws["C1"].value == 3
        assert ws["D1"].value == 4
        assert ws["E1"].value == 5
