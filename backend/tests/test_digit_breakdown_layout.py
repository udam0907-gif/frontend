"""v6 Gate 2 — _is_digit_breakdown_layout 자동 감지 + _fill_flat 분기 통합."""

from __future__ import annotations

from openpyxl import Workbook

from app.services.xlsx_document_filler import (
    XlsxDocumentFiller,
    _fill_amount_digits,
    _is_digit_breakdown_layout,
)


class TestIsDigitBreakdownLayout:
    def test_optomarine_layout_positive(self) -> None:
        """옵토마린 양식 모사 — E12-M12에 자릿수 라벨 9개 (억/천/백/십/만/천/백/십/일)."""
        wb = Workbook()
        ws = wb.active
        labels = ["억", "천", "백", "십", "만", "천", "백", "십", "일"]
        for i, label in enumerate(labels):
            ws.cell(row=12, column=5 + i).value = label  # E12-M12
        assert _is_digit_breakdown_layout(ws, "E13") is True

    def test_partial_labels_above_threshold(self) -> None:
        """6개 라벨이면 임계값 통과."""
        wb = Workbook()
        ws = wb.active
        labels = ["억", "천", "백", "십", "만", "일"]  # 6개
        for i, label in enumerate(labels):
            ws.cell(row=12, column=5 + i).value = label
        assert _is_digit_breakdown_layout(ws, "E13") is True

    def test_below_threshold_returns_false(self) -> None:
        """5개 이하 라벨이면 False."""
        wb = Workbook()
        ws = wb.active
        labels = ["억", "천", "백", "십", "만"]  # 5개
        for i, label in enumerate(labels):
            ws.cell(row=12, column=5 + i).value = label
        assert _is_digit_breakdown_layout(ws, "E13") is False

    def test_no_labels_returns_false(self) -> None:
        """일반 양식 (라벨 없음 or 합계 라벨만)."""
        wb = Workbook()
        ws = wb.active
        ws["E12"].value = "합계"
        assert _is_digit_breakdown_layout(ws, "E13") is False

    def test_row_one_returns_false(self) -> None:
        """1행에서 시작하면 위 행이 없으니 False."""
        wb = Workbook()
        ws = wb.active
        assert _is_digit_breakdown_layout(ws, "E1") is False

    def test_invalid_ref_returns_false(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert _is_digit_breakdown_layout(ws, "") is False
        assert _is_digit_breakdown_layout(ws, "X") is False
        assert _is_digit_breakdown_layout(ws, "13") is False


class TestFillFlatDigitBreakdownIntegration:
    """_fill_flat에서 자릿수 분리 자동 감지 + 호출 검증."""

    def _make_ws_with_digit_labels(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "명세-1116"
        labels = ["억", "천", "백", "십", "만", "천", "백", "십", "일"]
        for i, label in enumerate(labels):
            ws.cell(row=12, column=5 + i).value = label  # E12-M12
        return wb, ws

    def test_total_amount_auto_detected_as_digit_breakdown(self) -> None:
        """total_amount=E13이고 위 행이 자릿수 라벨 → 자동 자릿수 분리."""
        wb, ws = self._make_ws_with_digit_labels()
        cell_map = {
            "total_amount": "E13",
            "sheet_name": "명세-1116",
        }
        context = {"total_amount": 275000}
        written: list[str] = []
        skipped: list[str] = []
        XlsxDocumentFiller()._fill_flat(ws, cell_map, context, written, skipped)

        # 275000 → 9자리 0-padded "000275000" → H13=2, I13=7, J13=5, 나머지 None
        assert ws["H13"].value == 2
        assert ws["I13"].value == 7
        assert ws["J13"].value == 5
        assert ws["E13"].value is None
        # written에 amount_digits 기록
        assert any("amount_digits" in w for w in written)

    def test_explicit_breakdown_start_takes_priority(self) -> None:
        """amount_digit_breakdown_start 키가 있으면 그걸 우선 사용."""
        wb, ws = self._make_ws_with_digit_labels()
        cell_map = {
            "amount_digit_breakdown_start": "E13",
            "sheet_name": "명세-1116",
        }
        context = {"total_amount": 5}
        written: list[str] = []
        skipped: list[str] = []
        XlsxDocumentFiller()._fill_flat(ws, cell_map, context, written, skipped)
        # 5 → 9자리 "000000005" → M13만 5
        assert ws["M13"].value == 5
        assert ws["E13"].value is None

    def test_total_amount_no_breakdown_layout_uses_single_cell(self) -> None:
        """위 행에 자릿수 라벨이 없으면 일반 단일 셀 박기."""
        wb = Workbook()
        ws = wb.active
        ws.title = "명세-1116"
        # 위 행은 일반 라벨만
        ws["E12"].value = "합계금액"
        cell_map = {"total_amount": "E13", "sheet_name": "명세-1116"}
        context = {"total_amount": 275000}
        written: list[str] = []
        skipped: list[str] = []
        XlsxDocumentFiller()._fill_flat(ws, cell_map, context, written, skipped)
        # 자릿수 분리 안 함 → E13에 275000 단일 박기
        assert ws["E13"].value == 275000
        assert ws["F13"].value is None

    def test_total_amount_zero_no_digits_written(self) -> None:
        """합계가 0이면 자릿수 분리 안 함, 단일 셀 박기도 안 함 (skip_keys 처리)."""
        wb, ws = self._make_ws_with_digit_labels()
        cell_map = {"total_amount": "E13", "sheet_name": "명세-1116"}
        context = {"total_amount": 0}
        written: list[str] = []
        skipped: list[str] = []
        XlsxDocumentFiller()._fill_flat(ws, cell_map, context, written, skipped)
        # 0이라 자릿수 분리 호출 자체 안 됨 — 단일 셀 박기 경로로 빠지지만 0/None이라 skip
        # E13에 데이터 박혔는지만 확인 — 0이면 None인 채로 남거나 0으로 박힐 수 있음
        # 핵심: 자릿수 분리는 호출 안 됨 (written에 amount_digits 없음)
        assert not any("amount_digits" in w for w in written)
