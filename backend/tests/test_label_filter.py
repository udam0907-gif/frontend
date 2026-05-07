"""Gate 1 단위 테스트 — cell_map 라벨 셀 보호 가드."""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.xlsx_cell_mapper import (
    _is_label_cell,
    filter_label_cell_mappings,
)


class TestIsLabelCell:
    def test_korean_short_labels(self) -> None:
        assert _is_label_cell("년") is True
        assert _is_label_cell("월") is True
        assert _is_label_cell("일") is True
        assert _is_label_cell("억") is True
        assert _is_label_cell("천") is True
        assert _is_label_cell("백") is True

    def test_korean_phrase_labels_with_spaces(self) -> None:
        assert _is_label_cell("상     호") is True
        assert _is_label_cell("등록 번호") is True
        assert _is_label_cell("대 표 자") is True
        assert _is_label_cell("합 계 금 액") is True
        assert _is_label_cell("공 급 받 는 자") is True

    def test_english_total_label(self) -> None:
        assert _is_label_cell("TOTAL") is True
        assert _is_label_cell("Total") is True

    def test_data_values_are_not_labels(self) -> None:
        assert _is_label_cell("2026") is False
        assert _is_label_cell("㈜옵토마린") is False
        assert _is_label_cell("주식회사 유담") is False
        assert _is_label_cell("커피커피커피") is False
        assert _is_label_cell("123-45-67890") is False

    def test_empty_or_none(self) -> None:
        assert _is_label_cell("") is False
        assert _is_label_cell(None) is False
        assert _is_label_cell("   ") is False

    def test_non_string_returns_false(self) -> None:
        assert _is_label_cell(0) is False
        assert _is_label_cell(123) is False
        assert _is_label_cell(["상호"]) is False


class TestFilterLabelCellMappings:
    def _make_workbook_with_labels(self) -> str:
        """O5에 '상호' 라벨, Q5에 빈 값 / U8에 '종목' 라벨, W8에 빈 값 시트 생성."""
        wb = Workbook()
        ws = wb.active
        ws.title = "명세-1116"
        ws["O5"] = "상     호"
        ws["Q5"] = None  # 동적 입력 자리 (값 셀)
        ws["U8"] = "종     목"
        ws["W8"] = None
        ws["B12"] = "년"
        ws["C12"] = "월"
        ws["D12"] = "일"
        ws["B21"] = "TOTAL"
        ws["N21"] = None
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb.save(tmp.name)
        return tmp.name

    def test_removes_label_cell_mappings(self) -> None:
        path = self._make_workbook_with_labels()
        try:
            cell_map = {
                "sheet_name": "명세-1116",
                "recipient_name": "O5",            # 라벨 → 제거 대상
                "recipient_company_name": "Q5",    # 값 → 보존
                "recipient_business_item": "U8",   # 라벨 → 제거 대상
                "issue_date_year": "B12",          # 라벨 → 제거 대상
                "total_amount": "B21",             # 라벨 → 제거 대상
                "_meta": {"items_table": {"start_row": 15}},
            }
            filtered, removed = filter_label_cell_mappings(
                cell_map, path, sheet_name="명세-1116"
            )
            assert "recipient_name" not in filtered
            assert "recipient_business_item" not in filtered
            assert "issue_date_year" not in filtered
            assert "total_amount" not in filtered
            assert filtered.get("recipient_company_name") == "Q5"
            assert filtered.get("sheet_name") == "명세-1116"
            assert "_meta" in filtered
            assert len(removed) == 4
        finally:
            Path(path).unlink(missing_ok=True)

    def test_empty_input(self) -> None:
        filtered, removed = filter_label_cell_mappings({}, "", None)
        assert filtered == {}
        assert removed == []

    def test_unreadable_file_returns_unchanged(self) -> None:
        cell_map = {"recipient_name": "O5"}
        filtered, removed = filter_label_cell_mappings(
            cell_map, "/nonexistent/path.xlsx", None
        )
        assert filtered == cell_map
        assert removed == []

    def test_preserves_meta_keys(self) -> None:
        path = self._make_workbook_with_labels()
        try:
            cell_map = {
                "_meta": {"items_table": {"start_row": 15}},
                "_cell_map": {"nested": "x"},
                "sheet_name": "명세-1116",
            }
            filtered, _ = filter_label_cell_mappings(
                cell_map, path, sheet_name="명세-1116"
            )
            assert filtered == cell_map
        finally:
            Path(path).unlink(missing_ok=True)
