"""Gate 2 단위 테스트 — placeholder 직접 스캔 + cell_map 병합."""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook

from app.services.xlsx_cell_mapper import (
    PLACEHOLDER_TO_KEY,
    merge_placeholder_into_cell_map,
    scan_placeholder_cells,
)


def _make_optomarine_v2_like_xlsx() -> str:
    """옵토마린 v2 거래명세서 양식 모사 — 8개 placeholder 포함."""
    wb = Workbook()
    ws = wb.active
    ws.title = "명세-1116"
    ws["O4"] = "등록번호"
    ws["Q4"] = "(공급받는자 등록번호)"
    ws["O5"] = "상     호"
    ws["Q5"] = "(공급받는자 상호)"
    ws["O6"] = "대표자"
    ws["Q6"] = "(공급받는자 대표자)"
    ws["O7"] = "주     소"
    ws["Q7"] = "(공급받는자 주소)"
    ws["O8"] = "업     태"
    ws["Q8"] = "(공급받는자 업태)"
    ws["U8"] = "종     목"
    ws["W8"] = "(공급받는자 종목)"
    ws["O9"] = "메     일"
    ws["Q9"] = "(공급받는자 메일)"
    ws["O10"] = "전     화"
    ws["Q10"] = "(공급받는자 전화)"
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


class TestScanPlaceholderCells:
    def test_scans_all_eight_recipient_placeholders(self) -> None:
        path = _make_optomarine_v2_like_xlsx()
        try:
            result = scan_placeholder_cells(path, sheet_name="명세-1116")
            assert result.get("recipient_business_number") == "Q4"
            assert result.get("recipient_company_name") == "Q5"
            assert result.get("recipient_representative") == "Q6"
            assert result.get("recipient_address") == "Q7"
            assert result.get("recipient_business_type") == "Q8"
            assert result.get("recipient_business_item") == "W8"
            assert result.get("recipient_email") == "Q9"
            assert result.get("recipient_phone") == "Q10"
            assert len(result) == 8
        finally:
            Path(path).unlink(missing_ok=True)

    def test_unknown_sheet_falls_back_to_active(self) -> None:
        path = _make_optomarine_v2_like_xlsx()
        try:
            result = scan_placeholder_cells(path, sheet_name="존재안함")
            assert result.get("recipient_email") == "Q9"
        finally:
            Path(path).unlink(missing_ok=True)

    def test_empty_template_path_returns_empty(self) -> None:
        assert scan_placeholder_cells("", None) == {}

    def test_unreadable_file_returns_empty(self) -> None:
        assert scan_placeholder_cells("/nonexistent/path.xlsx", None) == {}

    def test_no_placeholders_returns_empty(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "그냥 글자"
        ws["A2"] = "100"
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb.save(tmp.name)
        try:
            assert scan_placeholder_cells(tmp.name) == {}
        finally:
            Path(tmp.name).unlink(missing_ok=True)

    def test_supplier_placeholders_also_detected(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "(공급자 상호)"
        ws["A2"] = "(공급자 대표자)"
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb.save(tmp.name)
        try:
            result = scan_placeholder_cells(tmp.name)
            assert result.get("supplier_company_name") == "A1"
            assert result.get("supplier_representative") == "A2"
        finally:
            Path(tmp.name).unlink(missing_ok=True)


class TestMergePlaceholderIntoCellMap:
    def test_adds_missing_keys(self) -> None:
        cell_map = {
            "recipient_company_name": "Q5",
            "sheet_name": "명세-1116",
        }
        placeholder_map = {
            "recipient_email": "Q9",
            "recipient_phone": "Q10",
        }
        merged, added = merge_placeholder_into_cell_map(cell_map, placeholder_map)
        assert merged["recipient_email"] == "Q9"
        assert merged["recipient_phone"] == "Q10"
        assert merged["recipient_company_name"] == "Q5"
        assert merged["sheet_name"] == "명세-1116"
        assert len(added) == 2

    def test_does_not_overwrite_existing_keys(self) -> None:
        """mapper 결과 우선 — 이미 있는 키는 placeholder 스캔이 덮지 않음."""
        cell_map = {"recipient_company_name": "Q5"}
        placeholder_map = {"recipient_company_name": "X99"}
        merged, added = merge_placeholder_into_cell_map(cell_map, placeholder_map)
        assert merged["recipient_company_name"] == "Q5"
        assert added == []

    def test_empty_cell_map(self) -> None:
        merged, added = merge_placeholder_into_cell_map(
            {}, {"recipient_email": "Q9"}
        )
        assert merged == {"recipient_email": "Q9"}
        assert added == ["recipient_email=Q9"]

    def test_empty_placeholder_map(self) -> None:
        cell_map = {"a": "A1"}
        merged, added = merge_placeholder_into_cell_map(cell_map, {})
        assert merged == cell_map
        assert added == []


class TestPlaceholderToKeyDict:
    def test_has_23_entries(self) -> None:
        # v6: 작성일 3개 + 자릿수 시작 1개 추가 (18 → 22)
        # v7: 받는자+귀하 결합 1개 추가 (22 → 23)
        assert len(PLACEHOLDER_TO_KEY) == 23

    def test_recipient_email_phone_present(self) -> None:
        assert PLACEHOLDER_TO_KEY["(공급받는자 메일)"] == "recipient_email"
        assert PLACEHOLDER_TO_KEY["(공급받는자 전화)"] == "recipient_phone"

    def test_v6_issue_date_keys_present(self) -> None:
        assert PLACEHOLDER_TO_KEY["(작성 년)"] == "issue_date_year"
        assert PLACEHOLDER_TO_KEY["(작성 월)"] == "issue_date_month"
        assert PLACEHOLDER_TO_KEY["(작성 일)"] == "issue_date_day"

    def test_v6_amount_digit_breakdown_key_present(self) -> None:
        assert (
            PLACEHOLDER_TO_KEY["(합계금액 자릿수 시작)"]
            == "amount_digit_breakdown_start"
        )

    def test_v7_recipient_with_honorific_key_present(self) -> None:
        assert "(공급받는자 상호+귀하)" in PLACEHOLDER_TO_KEY
        assert (
            PLACEHOLDER_TO_KEY["(공급받는자 상호+귀하)"]
            == "recipient_with_honorific"
        )
