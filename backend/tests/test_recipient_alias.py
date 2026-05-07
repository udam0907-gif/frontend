"""v4 Gate 3 — recipient_email/phone/fax alias 테스트."""

from __future__ import annotations

from openpyxl import Workbook

from app.services.xlsx_document_filler import XlsxDocumentFiller


class TestFieldAliasesDict:
    def test_recipient_email_has_our_company_fallback(self) -> None:
        aliases = XlsxDocumentFiller.FIELD_ALIASES["recipient_email"]
        assert "our_company_email" in aliases
        assert "company_email" in aliases

    def test_recipient_phone_has_our_company_fallback(self) -> None:
        aliases = XlsxDocumentFiller.FIELD_ALIASES["recipient_phone"]
        assert "our_company_phone" in aliases
        assert "company_phone" in aliases

    def test_recipient_fax_has_our_company_fallback(self) -> None:
        aliases = XlsxDocumentFiller.FIELD_ALIASES["recipient_fax"]
        assert "our_company_fax" in aliases
        assert "company_fax" in aliases


class TestFillFlatAliasResolution:
    def _setup_ws(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "명세-1116"
        return wb, ws

    def test_recipient_email_resolved_from_our_company_email(self) -> None:
        wb, ws = self._setup_ws()
        cell_map = {"recipient_email": "Q9", "sheet_name": "명세-1116"}
        context = {"our_company_email": "test@yudam.com"}
        written: list[str] = []
        skipped: list[str] = []
        XlsxDocumentFiller()._fill_flat(ws, cell_map, context, written, skipped)
        assert ws["Q9"].value == "test@yudam.com"
        assert "recipient_email=Q9" in written

    def test_recipient_phone_resolved_from_our_company_phone(self) -> None:
        wb, ws = self._setup_ws()
        cell_map = {"recipient_phone": "Q10", "sheet_name": "명세-1116"}
        context = {"our_company_phone": "010-1234-5678"}
        written: list[str] = []
        skipped: list[str] = []
        XlsxDocumentFiller()._fill_flat(ws, cell_map, context, written, skipped)
        assert ws["Q10"].value == "010-1234-5678"

    def test_direct_recipient_key_takes_priority_over_alias(self) -> None:
        """context에 recipient_email이 직접 있으면 그게 우선."""
        wb, ws = self._setup_ws()
        cell_map = {"recipient_email": "Q9", "sheet_name": "명세-1116"}
        context = {
            "recipient_email": "direct@x.com",
            "our_company_email": "fallback@x.com",
        }
        written: list[str] = []
        skipped: list[str] = []
        XlsxDocumentFiller()._fill_flat(ws, cell_map, context, written, skipped)
        assert ws["Q9"].value == "direct@x.com"

    def test_no_alias_no_value_skipped(self) -> None:
        """recipient_email 매핑 있는데 context에 메일 없으면 skip."""
        wb, ws = self._setup_ws()
        cell_map = {"recipient_email": "Q9", "sheet_name": "명세-1116"}
        context: dict = {}
        written: list[str] = []
        skipped: list[str] = []
        XlsxDocumentFiller()._fill_flat(ws, cell_map, context, written, skipped)
        assert ws["Q9"].value is None
        assert any("recipient_email" in s for s in skipped)

    def test_recipient_fax_alias(self) -> None:
        wb, ws = self._setup_ws()
        cell_map = {"recipient_fax": "Q11", "sheet_name": "명세-1116"}
        context = {"our_company_fax": "02-9876-5432"}
        written: list[str] = []
        skipped: list[str] = []
        XlsxDocumentFiller()._fill_flat(ws, cell_map, context, written, skipped)
        assert ws["Q11"].value == "02-9876-5432"
