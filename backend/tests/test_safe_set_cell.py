"""v4 Gate 1 — _safe_set_cell 라벨 가드 단위 테스트."""

from __future__ import annotations

from openpyxl import Workbook

from app.services.xlsx_document_filler import _safe_set_cell


class TestSafeSetCellLabelGuard:
    def test_label_cell_protected(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws["B21"] = "TOTAL"
        skip_log: list[str] = []
        result = _safe_set_cell(ws, "B21", 275000, label_skip_log=skip_log)
        assert result is False
        assert ws["B21"].value == "TOTAL"
        assert len(skip_log) == 1
        assert "B21" in skip_log[0]
        assert "label" in skip_log[0]

    def test_korean_label_year_protected(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws["B12"] = "년"
        result = _safe_set_cell(ws, "B12", 2026)
        assert result is False
        assert ws["B12"].value == "년"

    def test_value_cell_written_when_empty(self) -> None:
        wb = Workbook()
        ws = wb.active
        skip_log: list[str] = []
        result = _safe_set_cell(ws, "Q5", "유담", label_skip_log=skip_log)
        assert result is True
        assert ws["Q5"].value == "유담"
        assert skip_log == []

    def test_overwrites_non_label_data(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws["Q4"] = "881-56-00629-old"
        result = _safe_set_cell(ws, "Q4", "881-56-00629")
        assert result is True
        assert ws["Q4"].value == "881-56-00629"

    def test_skip_log_optional(self) -> None:
        """label_skip_log=None이어도 동작 — 호출 측이 로그 안 받아도 OK."""
        wb = Workbook()
        ws = wb.active
        ws["B21"] = "TOTAL"
        result = _safe_set_cell(ws, "B21", 275000)
        assert result is False
        assert ws["B21"].value == "TOTAL"

    def test_numeric_existing_value_not_treated_as_label(self) -> None:
        """기존 값이 숫자(예: 양식의 0)면 덮어쓰기 허용."""
        wb = Workbook()
        ws = wb.active
        ws["N15"] = 0
        result = _safe_set_cell(ws, "N15", 150000)
        assert result is True
        assert ws["N15"].value == 150000


class TestSafeSetCellFormulaGuard:
    def test_sum_formula_protected(self) -> None:
        """양식의 자동합계 수식(=SUM(...))은 덮어쓰지 않음."""
        wb = Workbook()
        ws = wb.active
        ws["N21"] = "=SUM(N15:Q20)"
        formula_log: list[str] = []
        result = _safe_set_cell(ws, "N21", 275000, formula_skip_log=formula_log)
        assert result is False
        assert ws["N21"].value == "=SUM(N15:Q20)"
        assert len(formula_log) == 1
        assert "N21" in formula_log[0]
        assert "formula" in formula_log[0]

    def test_multiply_formula_protected(self) -> None:
        """라인 수식(=J15*L15)도 보존."""
        wb = Workbook()
        ws = wb.active
        ws["N15"] = "=J15*L15"
        result = _safe_set_cell(ws, "N15", 150000)
        assert result is False
        assert ws["N15"].value == "=J15*L15"

    def test_non_formula_string_overwritten(self) -> None:
        """단순 텍스트는 라벨이 아닌 한 덮어씀 허용."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "old text"
        result = _safe_set_cell(ws, "A1", "new text")
        assert result is True
        assert ws["A1"].value == "new text"

    def test_label_takes_priority_over_formula(self) -> None:
        """라벨 가드가 수식 가드보다 먼저 — 둘 다면 label 로그만."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "TOTAL"
        label_log: list[str] = []
        formula_log: list[str] = []
        result = _safe_set_cell(
            ws, "A1", 100, label_skip_log=label_log, formula_skip_log=formula_log
        )
        assert result is False
        assert len(label_log) == 1
        assert len(formula_log) == 0

    def test_both_logs_optional(self) -> None:
        """formula_skip_log=None이어도 정상 거부."""
        wb = Workbook()
        ws = wb.active
        ws["N15"] = "=J15*L15"
        result = _safe_set_cell(ws, "N15", 999)
        assert result is False
        assert ws["N15"].value == "=J15*L15"
