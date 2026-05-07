"""v5 Gate 1 — 비주력 시트 보존 단위 테스트.

이전 동작(삭제) → 신규 동작(보존 + 활성 포인터만 변경) 회귀 방지.
실제 fill() 통합 테스트로 검증 (file I/O 포함).
"""

from __future__ import annotations

import shutil
import tempfile
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook

from app.services.xlsx_document_filler import XlsxDocumentFiller


def _make_two_sheet_template() -> str:
    """원본 템플릿 모사 — '자체판매'(첫 시트) + '명세-1116'(두 번째 시트)."""
    wb = Workbook()
    # 첫 시트: 자체판매 (1490행 시뮬레이션 — 데이터 잔존 시뮬)
    ws_sales = wb.active
    ws_sales.title = "자체판매"
    ws_sales["A1"] = "vendor 자체 판매 데이터 — 보존 대상"
    ws_sales["A100"] = "행 100 데이터"
    # 두 번째 시트: 명세-1116 (거래명세서)
    ws_main = wb.create_sheet("명세-1116")
    ws_main["B21"] = "TOTAL"  # 라벨
    ws_main["Q5"] = None       # 데이터 자리
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


class TestSecondarySheetsPreserved:
    @pytest.fixture
    def filler_with_tmp_storage(self, tmp_path):
        """settings.storage_documents_path를 임시 디렉토리로 패치."""
        with patch(
            "app.services.xlsx_document_filler.settings"
        ) as mock_settings:
            mock_settings.storage_documents_path = str(tmp_path)
            yield XlsxDocumentFiller()

    def test_all_sheets_remain_after_fill(self, filler_with_tmp_storage) -> None:
        template = _make_two_sheet_template()
        try:
            field_map = {
                "_cell_map": {
                    "sheet_name": "명세-1116",
                    "recipient_company_name": "Q5",
                }
            }
            context = {"recipient_company_name": "유담"}
            output = filler_with_tmp_storage.fill(
                template_path=template,
                field_map=field_map,
                context=context,
                expense_item_id="test-001",
            )
            wb = load_workbook(output)
            try:
                # 시트 2개 모두 보존
                assert "자체판매" in wb.sheetnames
                assert "명세-1116" in wb.sheetnames
                assert len(wb.sheetnames) == 2
                # 자체판매 데이터 보존
                assert wb["자체판매"]["A1"].value == "vendor 자체 판매 데이터 — 보존 대상"
                assert wb["자체판매"]["A100"].value == "행 100 데이터"
                # 명세-1116에 데이터 박힘
                assert wb["명세-1116"]["Q5"].value == "유담"
            finally:
                wb.close()
        finally:
            Path(template).unlink(missing_ok=True)

    def test_active_sheet_set_to_target(self, filler_with_tmp_storage) -> None:
        """파일 열면 거래명세서 시트가 먼저 보여야 함."""
        template = _make_two_sheet_template()
        try:
            field_map = {
                "_cell_map": {"sheet_name": "명세-1116", "recipient_company_name": "Q5"}
            }
            context = {"recipient_company_name": "유담"}
            output = filler_with_tmp_storage.fill(
                template_path=template,
                field_map=field_map,
                context=context,
                expense_item_id="test-002",
            )
            wb = load_workbook(output)
            try:
                # 파일 열었을 때 활성 시트가 명세-1116
                assert wb.active.title == "명세-1116"
            finally:
                wb.close()
        finally:
            Path(template).unlink(missing_ok=True)

    def test_single_sheet_template_unaffected(self, filler_with_tmp_storage) -> None:
        """시트 1개짜리 양식은 영향 없음 (회귀 점검)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "명세-1116"
        ws["Q5"] = None
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb.save(tmp.name)
        try:
            field_map = {
                "_cell_map": {"sheet_name": "명세-1116", "recipient_company_name": "Q5"}
            }
            context = {"recipient_company_name": "유담"}
            output = filler_with_tmp_storage.fill(
                template_path=tmp.name,
                field_map=field_map,
                context=context,
                expense_item_id="test-003",
            )
            out_wb = load_workbook(output)
            try:
                assert len(out_wb.sheetnames) == 1
                assert out_wb["명세-1116"]["Q5"].value == "유담"
            finally:
                out_wb.close()
        finally:
            Path(tmp.name).unlink(missing_ok=True)
