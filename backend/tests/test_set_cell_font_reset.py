"""Gate 3 단위 테스트 — _set_cell 폰트 리셋 (색만 검정, family/size/bold/italic 보존)."""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.services.xlsx_document_filler import _set_cell


def _make_wb_with_purple_font_cell() -> str:
    """A1에 보라색 굵은 라벨 폰트 셀 생성 → 파일 경로 반환."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "기존 라벨"
    ws["A1"].font = Font(
        name="맑은 고딕",
        size=14,
        bold=True,
        italic=False,
        color="FF6600FF",
    )
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


class TestSetCell:
    def test_color_resets_to_black_family_size_bold_preserved(self) -> None:
        path = _make_wb_with_purple_font_cell()
        try:
            wb = load_workbook(path)
            ws = wb.active

            _set_cell(ws, "A1", "새 데이터")

            cell = ws["A1"]
            assert cell.value == "새 데이터"
            # 색은 검정으로 리셋
            assert cell.font.color is not None
            assert cell.font.color.rgb == "FF000000"
            # family / size / bold / italic 보존
            assert cell.font.name == "맑은 고딕"
            assert cell.font.size == 14
            assert cell.font.bold is True
            assert cell.font.italic is False
            wb.close()
        finally:
            Path(path).unlink(missing_ok=True)

    def test_writes_value_when_cell_has_no_font(self) -> None:
        wb = Workbook()
        ws = wb.active
        _set_cell(ws, "B2", 12345)
        assert ws["B2"].value == 12345

    def test_int_value(self) -> None:
        wb = Workbook()
        ws = wb.active
        _set_cell(ws, "C3", 100)
        assert ws["C3"].value == 100

    def test_overwrites_existing_value(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws["D4"] = "원래 값"
        _set_cell(ws, "D4", "새 값")
        assert ws["D4"].value == "새 값"
