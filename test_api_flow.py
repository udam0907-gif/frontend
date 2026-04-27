"""
API 기반 전체 흐름 검증:
1. registry 조회
2. 템플릿 2개 업로드 (지출결의서, 견적서)
3. 셀 매핑 저장
4. 저장 후 field_map pre-fill 확인
5. 배지용 매핑 카운트 확인
6. 문서 생성 후 실제 셀 값 입력 확인
"""

import sys, uuid, requests, json
from pathlib import Path
import openpyxl

BASE = "http://localhost:8000/api/v1"
TEMPLATE_DIR = Path("test_templates")
OUTPUT_DIR   = Path("test_outputs")

# ── 테스트용 XLSX 생성 (없으면 재생성) ───────────────────────────────────────
def ensure_templates():
    TEMPLATE_DIR.mkdir(exist_ok=True)
    for name in ("expense_resolution_template.xlsx", "quote_template.xlsx"):
        p = TEMPLATE_DIR / name
        if not p.exists():
            wb = openpyxl.Workbook(); ws = wb.active
            ws["A1"] = name.replace("_template.xlsx","").replace("_"," ").title()
            wb.save(p)

ensure_templates()

CELL_MAPPING = {
    "company_name":   "B3",
    "execution_date": "E3",
    "budget_item":    "B5",
    "note":           "E5",
    "item_name":      "A9",
    "quantity":       "C9",
    "unit_price":     "D9",
    "amount":         "E9",
}

# ── 1. registry 조회 ──────────────────────────────────────────────────────────
print("\n[1] GET /templates/fields/registry")
r = requests.get(f"{BASE}/templates/fields/registry")
assert r.status_code == 200, f"FAIL: {r.status_code} {r.text}"
fields = r.json()["fields"]
print(f"  OK — {len(fields)}개 필드")

# ── 2. 프로젝트 조회 (첫 번째) ────────────────────────────────────────────────
print("\n[2] GET /projects (첫 번째 프로젝트 사용)")
r = requests.get(f"{BASE}/projects")
assert r.status_code == 200 and r.json(), "프로젝트 없음"
project_id = r.json()[0]["id"]
print(f"  project_id={project_id}")

# ── 3. 템플릿 2개 업로드 ──────────────────────────────────────────────────────
def upload_template(filename: str, doc_type: str, name: str) -> str:
    with open(TEMPLATE_DIR / filename, "rb") as f:
        r = requests.post(f"{BASE}/templates", data={
            "name": name,
            "category_type": "materials",
            "document_type": doc_type,
        }, files={"file": (filename, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 201, f"업로드 실패: {r.status_code} {r.text}"
    return r.json()["id"]

print("\n[3] 템플릿 업로드")
exp_id   = upload_template("expense_resolution_template.xlsx", "expense_resolution", "지출결의서 테스트")
quote_id = upload_template("quote_template.xlsx",              "quote",              "견적서 테스트")
print(f"  지출결의서 template_id={exp_id}")
print(f"  견적서     template_id={quote_id}")

# ── 4. 셀 매핑 저장 ───────────────────────────────────────────────────────────
print("\n[4] PUT /templates/{id}/cell-mapping")
for tid, label in [(exp_id, "지출결의서"), (quote_id, "견적서")]:
    r = requests.put(f"{BASE}/templates/{tid}/cell-mapping", json={"mapping": CELL_MAPPING})
    assert r.status_code == 200, f"{label} 매핑 저장 실패: {r.status_code} {r.text}"
    fm = r.json()["field_map"]
    saved_cells = {k: v["cell"] for k, v in fm.items() if isinstance(v, dict) and v.get("cell")}
    print(f"  {label}: {len(saved_cells)}개 저장 — {saved_cells}")

# ── 5. pre-fill 확인 (GET 후 field_map.cell 검증) ────────────────────────────
print("\n[5] GET /templates/{id} — pre-fill 확인")
for tid, label in [(exp_id, "지출결의서"), (quote_id, "견적서")]:
    r = requests.get(f"{BASE}/templates/{tid}")
    assert r.status_code == 200
    fm = r.json()["field_map"]
    cells = {k: v["cell"] for k, v in fm.items() if isinstance(v, dict) and v.get("cell")}
    match = all(cells.get(k) == v for k, v in CELL_MAPPING.items())
    badge_count = len(cells)
    print(f"  {label}: pre-fill={'OK' if match else 'FAIL'}, 배지카운트={badge_count}")

# ── 6. 비용집행 생성 ──────────────────────────────────────────────────────────
print("\n[6] POST /expenses — 비용집행 생성")
r = requests.post(f"{BASE}/expenses", json={
    "project_id": project_id,
    "category_type": "materials",
    "title": "정밀측정장비 구매",
    "amount": 3000000,
    "vendor_name": "주식회사 테스트업체",
    "expense_date": "2026-04-22",
    "metadata": {
        "company_name":   "주식회사 테스트업체",
        "execution_date": "2026-04-22",
        "budget_item":    "재료비",
        "note":           "R&D 장비 구매",
        "item_name":      "정밀측정장비",
        "quantity":       2,
        "unit_price":     1500000,
        "amount":         3000000,
    }
})
assert r.status_code == 201, f"비용집행 생성 실패: {r.status_code} {r.text}"
expense_id = r.json()["id"]
print(f"  expense_id={expense_id}")

# ── 7. 문서 생성 ──────────────────────────────────────────────────────────────
print("\n[7] POST /documents/generate — 지출결의서 문서 생성")
meta = {
    "company_name":   "주식회사 테스트업체",
    "execution_date": "2026-04-22",
    "budget_item":    "재료비",
    "note":           "R&D 장비 구매",
    "item_name":      "정밀측정장비",
    "quantity":       2,
    "unit_price":     1500000,
    "amount":         3000000,
}
r = requests.post(f"{BASE}/documents/generate", json={
    "expense_item_id": expense_id,
    "template_id": exp_id,
    "field_values": meta,
})
assert r.status_code == 201, f"문서 생성 실패: {r.status_code} {r.text}"
doc = r.json()
doc_id = doc["id"]
render_mode = doc.get("generation_trace", {}).get("render_mode", "unknown")
print(f"  doc_id={doc_id}, render_mode={render_mode}")

# ── 8. 다운로드 후 셀 값 검증 ─────────────────────────────────────────────────
print("\n[8] GET /documents/{id}/download — 셀 값 검증")
r = requests.get(f"{BASE}/documents/{doc_id}/download")
assert r.status_code == 200, f"다운로드 실패: {r.status_code}"
out_path = OUTPUT_DIR / f"api_test_result_{doc_id[:8]}.xlsx"
OUTPUT_DIR.mkdir(exist_ok=True)
out_path.write_bytes(r.content)

wb = openpyxl.load_workbook(out_path)
ws = wb.active
ok, fail = [], []
for field, cell in CELL_MAPPING.items():
    actual = ws[cell].value
    expected = meta.get(field)
    if expected is not None:
        if str(actual) == str(expected):
            ok.append(f"  [OK] {field}({cell})={actual}")
        else:
            fail.append(f"  [FAIL] {field}({cell}): expected={expected}, actual={actual}")

for line in ok:   print(line)
for line in fail: print(line)

print(f"\n결과: {len(ok)}/{len(ok)+len(fail)} 성공, render_mode={render_mode}")
if not fail and render_mode == "excel_rendered":
    print("전체 통과")
else:
    print("주의: 일부 미통과")
