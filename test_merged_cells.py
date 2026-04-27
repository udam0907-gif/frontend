"""병합셀 구조 분석 — company_name/execution_date 셀 주소 실제 확인"""
import openpyxl
from pathlib import Path

BASE = Path("C:/Users/FORYOUCOM/cm_app/backend")

TARGET_CELLS = {"B3", "E3", "B5", "E5", "A9", "C9", "D9", "E9"}
CHECK_FIELDS  = {"B3": "company_name", "E3": "execution_date",
                 "B5": "budget_item",  "E5": "note"}

FILES = {
    "견적서(테스트)":        "storage/templates/ebe1f76e3dbd7e7d6a0724560fc3f7d4.xlsx",
    "지출결의서(유담1)":     "storage/templates/58a61d4057fe1d701c4f959359bdfa61.xlsx",
    "지출결의서(유담2)":     "storage/templates/0ec43020a0518ec66d19f27f82c081c6.xlsx",
    "검수확인서(재료비)":    "storage/templates/8fb57a2120d9293779bf519c8a982fc6.xlsx",
    "업체견적서(피닉)":      "storage/documents/vendors/339cbf90-e167-43d3-91db-1d11dd5c5f78/quote_template_피닉_거래명세서_양식.xlsx",
    "업체견적서(선양)":      "storage/documents/vendors/ef762400-688c-4d51-aa6c-d440e7283299/quote_template_선양견적서.xlsx",
}

def get_merge_master(ws, cell_addr: str) -> str:
    """셀이 병합 범위에 속하면 좌상단(master) 주소 반환, 아니면 원래 주소 반환."""
    for merge in ws.merged_cells.ranges:
        if cell_addr in [c.coordinate for row in merge.cells for c in row]:
            # min_row, min_col이 master
            import openpyxl.utils as utils
            master = f"{utils.get_column_letter(merge.min_col)}{merge.min_row}"
            return f"{cell_addr}→MERGED(master={master}, range={merge})"
    return f"{cell_addr}→단독셀"

print("=" * 70)
for label, rel_path in FILES.items():
    full = BASE / rel_path
    if not full.exists():
        print(f"\n[{label}] 파일 없음: {rel_path}")
        continue

    try:
        wb = openpyxl.load_workbook(full)
        ws = wb.active
    except Exception as e:
        print(f"\n[{label}] 열기 실패: {e}")
        continue

    print(f"\n[{label}]  (시트: {ws.title})")
    has_merge_issue = False
    for cell_addr in sorted(TARGET_CELLS):
        result = get_merge_master(ws, cell_addr)
        current_val = ws[cell_addr].value
        field = CHECK_FIELDS.get(cell_addr, "")
        flag = " ← 병합!" if "MERGED" in result else ""
        print(f"  {cell_addr}({field:15}) {result:45}  현재값={repr(current_val)}{flag}")
        if "MERGED" in result and cell_addr in CHECK_FIELDS:
            has_merge_issue = True

    if has_merge_issue:
        print(f"  *** 병합셀 문제 있음 — 위 master 주소로 매핑 교체 필요 ***")
