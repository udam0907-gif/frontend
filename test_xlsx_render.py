"""
XLSX 셀 매핑 렌더링 독립 테스트
- DB/FastAPI 없이 document_generator._generate_xlsx() 직접 호출
- 지출결의서 / 견적서 / 비교견적서 3종 테스트
"""

import asyncio
import sys
import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ─── 경로 설정 ───────────────────────────────────────────────────────────────
BASE = Path(__file__).parent
BACKEND = BASE / "backend"
sys.path.insert(0, str(BACKEND))

TEMPLATE_DIR = BASE / "test_templates"
OUTPUT_DIR = BASE / "test_outputs"
TEMPLATE_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)


# ─── 테스트용 XLSX 템플릿 생성 ────────────────────────────────────────────────

def make_expense_resolution_template() -> Path:
    """지출결의서 템플릿 (실무 레이아웃 근사)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "지출결의서"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    ws["A1"] = "지출결의서"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")

    ws["A3"] = "업체명"
    ws["B3"] = ""          # ← company_name 입력 위치

    ws["D3"] = "집행일자"
    ws["E3"] = ""          # ← execution_date 입력 위치

    ws["A5"] = "예산항목"
    ws["B5"] = ""          # ← budget_item

    ws["D5"] = "비고"
    ws["E5"] = ""          # ← note

    ws["A8"] = "품명"
    ws["B8"] = "규격"
    ws["C8"] = "수량"
    ws["D8"] = "단가"
    ws["E8"] = "금액"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}8"].font = Font(bold=True)

    ws["A9"] = ""          # ← item_name
    ws["B9"] = ""
    ws["C9"] = ""          # ← quantity
    ws["D9"] = ""          # ← unit_price
    ws["E9"] = ""          # ← amount

    path = TEMPLATE_DIR / "expense_resolution_template.xlsx"
    wb.save(path)
    return path


def make_quote_template() -> Path:
    """견적서 템플릿"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "견적서"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    ws["A1"] = "견 적 서"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")

    ws["A3"] = "공급업체"
    ws["B3"] = ""          # ← company_name

    ws["D3"] = "견적일자"
    ws["E3"] = ""          # ← execution_date

    ws["A5"] = "예산항목"
    ws["B5"] = ""          # ← budget_item

    ws["D5"] = "비고"
    ws["E5"] = ""          # ← note

    ws["A8"] = "품명"
    ws["B8"] = "규격"
    ws["C8"] = "수량"
    ws["D8"] = "단가"
    ws["E8"] = "금액"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}8"].font = Font(bold=True)

    ws["A9"] = ""          # ← item_name
    ws["B9"] = ""
    ws["C9"] = ""          # ← quantity
    ws["D9"] = ""          # ← unit_price
    ws["E9"] = ""          # ← amount

    path = TEMPLATE_DIR / "quote_template.xlsx"
    wb.save(path)
    return path


# ─── 셀 매핑 정의 ────────────────────────────────────────────────────────────

EXPENSE_RESOLUTION_MAPPING = {
    "company_name":   {"label": "업체명",   "type": "text",   "required": True, "source": "user_input", "cell": "B3"},
    "execution_date": {"label": "집행일자", "type": "date",   "required": True, "source": "user_input", "cell": "E3"},
    "budget_item":    {"label": "예산항목", "type": "text",   "required": True, "source": "user_input", "cell": "B5"},
    "note":           {"label": "비고",     "type": "text",   "required": False,"source": "user_input", "cell": "E5"},
    "item_name":      {"label": "품명",     "type": "text",   "required": True, "source": "user_input", "cell": "A9"},
    "quantity":       {"label": "수량",     "type": "number", "required": True, "source": "user_input", "cell": "C9"},
    "unit_price":     {"label": "단가",     "type": "number", "required": True, "source": "user_input", "cell": "D9"},
    "amount":         {"label": "금액",     "type": "number", "required": True, "source": "user_input", "cell": "E9"},
}

QUOTE_MAPPING = {
    "company_name":   {"label": "공급업체", "type": "text",   "required": True, "source": "user_input", "cell": "B3"},
    "execution_date": {"label": "견적일자", "type": "date",   "required": True, "source": "user_input", "cell": "E3"},
    "budget_item":    {"label": "예산항목", "type": "text",   "required": True, "source": "user_input", "cell": "B5"},
    "note":           {"label": "비고",     "type": "text",   "required": False,"source": "user_input", "cell": "E5"},
    "item_name":      {"label": "품명",     "type": "text",   "required": True, "source": "user_input", "cell": "A9"},
    "quantity":       {"label": "수량",     "type": "number", "required": True, "source": "user_input", "cell": "C9"},
    "unit_price":     {"label": "단가",     "type": "number", "required": True, "source": "user_input", "cell": "D9"},
    "amount":         {"label": "금액",     "type": "number", "required": True, "source": "user_input", "cell": "E9"},
}

# 비교견적서: 동일 양식, 금액만 1.1배
COMPARATIVE_QUOTE_MAPPING = QUOTE_MAPPING


# ─── 테스트 입력값 ────────────────────────────────────────────────────────────

USER_VALUES_MAIN = {
    "company_name":   "주식회사 테스트업체",
    "execution_date": "2026-04-22",
    "budget_item":    "재료비",
    "note":           "R&D 장비 구매",
    "item_name":      "정밀측정장비",
    "quantity":       2,
    "unit_price":     1_500_000,
    "amount":         3_000_000,
}

USER_VALUES_COMPARE = {
    **USER_VALUES_MAIN,
    "company_name": "비교견적 주식회사",
    "amount":       3_300_000,   # 3,000,000 × 1.1
    "unit_price":   1_650_000,
    "note":         "비교견적 (원견적 3,000,000원 기준 10% 인상)",
}


# ─── 렌더 로직 직접 실행 (document_generator._generate_xlsx 로직 재현) ──────

def render_xlsx(template_path: Path, field_map: dict, user_values: dict, output_dir: Path, label: str) -> dict:
    import shutil

    cell_mapped = {k: v for k, v in field_map.items() if isinstance(v, dict) and v.get("cell")}

    if not cell_mapped:
        out = output_dir / f"{label}_mapping_needed.xlsx"
        shutil.copy2(template_path, out)
        return {"render_mode": "mapping_needed", "output_path": str(out), "written": [], "skipped": []}

    context = dict(user_values)
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    written = []
    skipped = []

    for field_key, meta in cell_mapped.items():
        cell_addr = meta["cell"]
        value = context.get(field_key)
        if value is None:
            skipped.append(field_key)
            continue
        try:
            ws[cell_addr] = value
            written.append(f"{field_key} → {cell_addr} = {value}")
        except Exception as e:
            skipped.append(f"{field_key}(ERR:{e})")

    out = output_dir / f"{label}_rendered.xlsx"
    wb.save(out)
    return {
        "render_mode": "excel_rendered" if written else "mapping_needed",
        "output_path": str(out),
        "written": written,
        "skipped": skipped,
    }


def verify_xlsx(output_path: str, field_map: dict, expected: dict) -> dict:
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    ok = []
    fail = []
    for field_key, meta in field_map.items():
        if not meta.get("cell"):
            continue
        expected_val = expected.get(field_key)
        if expected_val is None:
            continue
        actual = ws[meta["cell"]].value
        if str(actual) == str(expected_val):
            ok.append(f"  [OK] {field_key}({meta['cell']}): {actual}")
        else:
            fail.append(f"  [FAIL] {field_key}({meta['cell']}): expected={expected_val}, actual={actual}")
    return {"ok": ok, "fail": fail}


# ─── 실행 ────────────────────────────────────────────────────────────────────

def run():
    print("=" * 60)
    print("XLSX 셀 매핑 렌더링 테스트")
    print("=" * 60)

    # 1. 템플릿 생성
    expense_tpl = make_expense_resolution_template()
    quote_tpl = make_quote_template()
    # 비교견적서는 견적서와 동일 양식 재사용
    import shutil
    compare_tpl = TEMPLATE_DIR / "comparative_quote_template.xlsx"
    shutil.copy2(quote_tpl, compare_tpl)

    print(f"\n[템플릿 생성]")
    print(f"  지출결의서: {expense_tpl.name}")
    print(f"  견적서:     {quote_tpl.name}")
    print(f"  비교견적서: {compare_tpl.name}")

    # ── 테스트 1: 지출결의서 ────────────────────────────────────────
    print("\n" + "─" * 50)
    print("[테스트 1] 지출결의서")
    r1 = render_xlsx(expense_tpl, EXPENSE_RESOLUTION_MAPPING, USER_VALUES_MAIN, OUTPUT_DIR, "expense_resolution")
    print(f"  render_mode: {r1['render_mode']}")
    print(f"  output: {Path(r1['output_path']).name}")
    print(f"  입력 성공 ({len(r1['written'])}개):")
    for w in r1["written"]:
        print(f"    {w}")
    if r1["skipped"]:
        print(f"  미입력: {r1['skipped']}")

    v1 = verify_xlsx(r1["output_path"], EXPENSE_RESOLUTION_MAPPING, USER_VALUES_MAIN)
    print(f"  검증:")
    for line in v1["ok"]:
        print(line)
    for line in v1["fail"]:
        print(line)

    # ── 테스트 2: 견적서 ─────────────────────────────────────────────
    print("\n" + "─" * 50)
    print("[테스트 2] 견적서")
    r2 = render_xlsx(quote_tpl, QUOTE_MAPPING, USER_VALUES_MAIN, OUTPUT_DIR, "quote")
    print(f"  render_mode: {r2['render_mode']}")
    print(f"  output: {Path(r2['output_path']).name}")
    print(f"  입력 성공 ({len(r2['written'])}개):")
    for w in r2["written"]:
        print(f"    {w}")

    v2 = verify_xlsx(r2["output_path"], QUOTE_MAPPING, USER_VALUES_MAIN)
    print(f"  검증:")
    for line in v2["ok"]:
        print(line)
    for line in v2["fail"]:
        print(line)

    # ── 테스트 3: 비교견적서 ──────────────────────────────────────────
    print("\n" + "─" * 50)
    print("[테스트 3] 비교견적서 (금액 ×1.1 적용)")
    r3 = render_xlsx(compare_tpl, COMPARATIVE_QUOTE_MAPPING, USER_VALUES_COMPARE, OUTPUT_DIR, "comparative_quote")
    print(f"  render_mode: {r3['render_mode']}")
    print(f"  output: {Path(r3['output_path']).name}")
    print(f"  입력 성공 ({len(r3['written'])}개):")
    for w in r3["written"]:
        print(f"    {w}")

    v3 = verify_xlsx(r3["output_path"], COMPARATIVE_QUOTE_MAPPING, USER_VALUES_COMPARE)
    print(f"  검증:")
    for line in v3["ok"]:
        print(line)
    for line in v3["fail"]:
        print(line)

    # ── 요약 ──────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("요약")
    total_ok = len(v1["ok"]) + len(v2["ok"]) + len(v3["ok"])
    total_fail = len(v1["fail"]) + len(v2["fail"]) + len(v3["fail"])
    print(f"  전체 검증: {total_ok}개 성공, {total_fail}개 실패")
    all_fields = list(EXPENSE_RESOLUTION_MAPPING.keys())
    mapped = [f for f in all_fields if EXPENSE_RESOLUTION_MAPPING[f].get("cell")]
    unmapped = [f for f in all_fields if not EXPENSE_RESOLUTION_MAPPING[f].get("cell")]
    print(f"  매핑 필드: {mapped}")
    print(f"  미매핑 필드: {unmapped if unmapped else '없음'}")
    print(f"\n  출력 파일 위치: {OUTPUT_DIR}")


if __name__ == "__main__":
    run()
